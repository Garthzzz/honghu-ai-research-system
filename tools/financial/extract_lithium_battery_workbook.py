from __future__ import annotations

"""Extract auditable quarterly operating series from the battery model workbook.

The workbook is used as a research seed, not as the live calculator engine.  We
therefore preserve cached values, formulas, section context and formula defects
in JSON while removing fixed-row dependencies from subsequent models.
"""

import argparse
import hashlib
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = next(ROOT.glob("*20260614*.xlsx"))
DEFAULT_OUTPUT = (
    ROOT
    / "cache"
    / "lithium_battery_research"
    / "models"
    / "battery_quarterly_workbook_extract_v1.json"
)

TARGETS = {
    "宁德时代 (调整)": {"company": "宁德时代", "ticker": "300750.SZ"},
    "鹏辉能源": {"company": "鹏辉能源", "ticker": "300438.SZ"},
    "比亚迪 (2)": {"company": "比亚迪", "ticker": "002594.SZ"},
    "亿纬锂能": {"company": "亿纬锂能", "ticker": "300014.SZ"},
    "国轩高科": {"company": "国轩高科", "ticker": "002074.SZ"},
    "欣旺达": {"company": "欣旺达", "ticker": "300207.SZ"},
    "孚能科技": {"company": "孚能科技", "ticker": "688567.SH"},
}

PERIOD_PATTERN = re.compile(
    r"^(?:[1-4]Q\d{2,4}E?|[12]H\d{2,4}E?|FY\d{2,4}E?|\d{4}[QH][1-4]E?|\d{4}AE?)$",
    re.I,
)
SECTION_PATTERN = re.compile(r"^\s*(?:[一二三四五六七八九十]+[、.]|\d+[、.])")
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!"}

INCLUDE_LABELS = re.compile(
    "|".join(
        (
            "收入",
            "销量",
            "销售",
            "出货",
            "产能",
            "产量",
            "装机",
            "均价",
            "成本",
            "毛利",
            "净利",
            "单瓦",
            "单wh",
            "单位利润",
            "费用",
            "研发",
            "减值",
            "投资收益",
            "现金流",
            "资本开支",
            "固定资产",
            "在建工程",
            "少数股东",
            "税率",
            "补贴",
            "其他收益",
            "公允价值",
            "海外",
            "铁锂",
            "三元",
            "储能",
            "动力",
            "消费",
        )
    ),
    re.I,
)


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return str(value)


def _sha256(path: Path) -> str:
    return "sha256:" + hashlib.sha256(path.read_bytes()).hexdigest()


def _normalise_period(value: Any) -> str | None:
    if hasattr(value, "year") and hasattr(value, "month"):
        quarter = (int(value.month) - 1) // 3 + 1
        return f"{quarter}Q{int(value.year)}"
    if not isinstance(value, str):
        return None
    text = re.sub(r"\s+", "", value).upper()
    if PERIOD_PATTERN.fullmatch(text):
        return text
    return None


def _find_header_row(sheet: Any) -> tuple[int, dict[int, str]]:
    best: tuple[int, int, dict[int, str]] | None = None
    for row in range(1, min(sheet.max_row, 12) + 1):
        periods = {
            column: period
            for column in range(1, sheet.max_column + 1)
            if (period := _normalise_period(sheet.cell(row, column).value))
        }
        candidate = (len(periods), row, periods)
        if best is None or candidate[0] > best[0]:
            best = candidate
    if best is None or best[0] < 4:
        raise ValueError(f"{sheet.title}: cannot locate quarterly header")
    return best[1], best[2]


def _find_label_column(sheet: Any, header_row: int) -> int:
    scores: list[tuple[int, int]] = []
    for column in range(1, min(8, sheet.max_column) + 1):
        score = 0
        for row in range(header_row + 1, sheet.max_row + 1):
            value = sheet.cell(row, column).value
            if isinstance(value, str) and value.strip():
                score += 1
        scores.append((score, column))
    return max(scores)[1]


def _extract_sheet(values: Any, formulas: Any, meta: dict[str, str]) -> dict[str, Any]:
    header_row, periods = _find_header_row(values)
    label_column = _find_label_column(values, header_row)
    rows: list[dict[str, Any]] = []
    formula_defects: list[dict[str, Any]] = []
    current_section = "公司总表"
    external_formula_count = 0
    formula_count = 0

    for row in range(header_row + 1, values.max_row + 1):
        raw_label = values.cell(row, label_column).value
        if not isinstance(raw_label, str) or not raw_label.strip():
            continue
        label = " ".join(raw_label.split())
        if SECTION_PATTERN.match(label):
            current_section = label
        if not INCLUDE_LABELS.search(label):
            continue

        observations: list[dict[str, Any]] = []
        for column, period in periods.items():
            cached = values.cell(row, column).value
            formula = formulas.cell(row, column).value
            if isinstance(formula, str) and formula.startswith("="):
                formula_count += 1
                if "[" in formula or "!" in formula:
                    external_formula_count += 1
                    formula_defects.append(
                        {
                            "cell": f"{get_column_letter(column)}{row}",
                            "period": period,
                            "label": label,
                            "issue": "external_or_cross_sheet_reference",
                            "formula": formula,
                        }
                    )
            if isinstance(cached, str) and cached in ERROR_VALUES:
                formula_defects.append(
                    {
                        "cell": f"{get_column_letter(column)}{row}",
                        "period": period,
                        "label": label,
                        "issue": "cached_formula_error",
                        "cached_value": cached,
                        "formula": formula if isinstance(formula, str) else None,
                    }
                )
            if cached is None and formula is None:
                continue
            observations.append(
                {
                    "period": period,
                    "value": _json_value(cached),
                    "formula": formula
                    if isinstance(formula, str) and formula.startswith("=")
                    else None,
                    "cell": f"{get_column_letter(column)}{row}",
                }
            )
        if observations:
            rows.append(
                {
                    "row": row,
                    "section": current_section,
                    "metric": label,
                    "observations": observations,
                }
            )

    return {
        **meta,
        "sheet_name": values.title,
        "header_row": header_row,
        "label_column": get_column_letter(label_column),
        "periods": list(periods.values()),
        "series": rows,
        "formula_audit": {
            "formula_count_in_selected_series": formula_count,
            "external_or_cross_sheet_formula_count": external_formula_count,
            "defect_count": len(formula_defects),
            "defects": formula_defects,
        },
    }


def extract(workbook: Path = DEFAULT_WORKBOOK) -> dict[str, Any]:
    # Normal mode is deliberate: openpyxl read_only random access is several
    # orders of magnitude slower for this workbook.
    value_book = load_workbook(workbook, data_only=True, read_only=False)
    formula_book = load_workbook(
        workbook, data_only=False, read_only=False, keep_links=True
    )
    companies = []
    for sheet_name, meta in TARGETS.items():
        if sheet_name not in value_book.sheetnames:
            raise KeyError(f"Missing required workbook sheet: {sheet_name}")
        companies.append(
            _extract_sheet(
                value_book[sheet_name],
                formula_book[sheet_name],
                meta,
            )
        )
    return {
        "schema_version": "lithium_battery.workbook_extract.v1",
        "extracted_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "workbook": workbook.name,
        "workbook_sha256": _sha256(workbook),
        "source_role": (
            "季度业务拆分与历史预测种子；外链公式、缓存错误和预测期均不自动视为事实。"
        ),
        "performance_note": (
            "采用非read-only顺序读取；该工作簿在read_only模式下随机取单元格会异常缓慢。"
        ),
        "companies": companies,
        "summary": {
            "company_count": len(companies),
            "series_count": sum(len(item["series"]) for item in companies),
            "formula_defect_count": sum(
                item["formula_audit"]["defect_count"] for item in companies
            ),
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    result = extract(args.workbook.resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(result["summary"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
