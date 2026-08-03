"""Recalculate the three researcher-supplied copper workbooks into a JSON snapshot.

The source workbooks contain a few explanatory cells beginning with ``=``.
Excel therefore treats those prose notes as malformed formulas and refuses to
open the files.  This utility never modifies the source workbooks.  It copies
cell values/formulas into a temporary workbook, converts only formula-looking
Chinese prose back to text, asks the locally installed Excel calculation engine
to rebuild formulas, and freezes the calculated tables with source SHA256.

The generated JSON is a deployable, read-only research input for the Flask
calculator.  The web calculator never needs Excel or COM on the broadcast VM.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = (
    ROOT
    / "config"
    / "copper_calculator_models"
    / "workbook_recalculation_snapshot_v1.json"
)
WORKBOOKS = {
    "紫金矿业": ROOT / "ZijinMining_601899.SH_Financial_Model_2026-07-17.xlsx",
    "洛阳钼业": ROOT / "CMOC_603993.SH_Financial_Model_2026-07-27.xlsx",
    "五矿资源": ROOT / "MMG_1208.HK_Financial_Model_2026-07-28.xlsx",
}
ANNUAL_SHEETS = (
    "Operating Drivers",
    "Revenue Model",
    "Income Statement",
    "Cash Flow",
)
YEARS = (2025, 2026, 2027, 2028, 2029, 2030)
CHINESE_RE = re.compile(r"[\u3400-\u9fff]")
PROSE_FORMULA_RE = re.compile(r"[×≈Σ]|(?:^|[=+\-*/(])\$\d")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return "sha256:" + digest.hexdigest()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _build_excel_readable_copy(source: Path, destination: Path) -> list[dict[str, str]]:
    old = load_workbook(source, data_only=False)
    new = Workbook()
    new.remove(new.active)
    repairs: list[dict[str, str]] = []
    for old_sheet in old.worksheets:
        sheet = new.create_sheet(old_sheet.title)
        for row in old_sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if (
                    isinstance(value, str)
                    and value.startswith("=")
                    and (
                        CHINESE_RE.search(value)
                        or PROSE_FORMULA_RE.search(value)
                    )
                ):
                    repairs.append(
                        {
                            "sheet": old_sheet.title,
                            "cell": cell.coordinate,
                            "text": value[1:],
                        }
                    )
                    value = value[1:]
                sheet.cell(cell.row, cell.column).value = value
    new.save(destination)
    return repairs


def _annual_table(sheet: Any) -> dict[str, Any]:
    headers = {
        str(_json_value(sheet.Cells(3, column).Value)): column
        for column in range(3, 15)
        if sheet.Cells(3, column).Value is not None
    }
    year_columns: dict[int, int] = {}
    for header, column in headers.items():
        match = re.search(r"(202[5-9]|2030)", header)
        if match:
            year_columns[int(match.group(1))] = column
    # The supplied models consistently place FY2025A-FY2030E in G:L.
    if not all(year in year_columns for year in YEARS):
        year_columns = {year: 7 + offset for offset, year in enumerate(YEARS)}
    rows: list[dict[str, Any]] = []
    for row_no in range(1, int(sheet.UsedRange.Rows.Count) + 1):
        label = sheet.Cells(row_no, 2).Value
        if label is None:
            continue
        values = {
            str(year): _json_value(sheet.Cells(row_no, column).Value)
            for year, column in year_columns.items()
        }
        if row_no <= 3 or all(value is None for value in values.values()):
            continue
        rows.append(
            {
                "row": row_no,
                "metric": str(label).strip(),
                "values": values,
            }
        )
    return {"years": list(YEARS), "rows": rows}


def _dcf_snapshot(sheet: Any) -> dict[str, Any]:
    summary: list[dict[str, Any]] = []
    annual: list[dict[str, Any]] = []
    for row_no in range(1, int(sheet.UsedRange.Rows.Count) + 1):
        label = sheet.Cells(row_no, 2).Value
        if label is None:
            continue
        label = str(label).strip()
        c_value = sheet.Cells(row_no, 3).Value
        if c_value is not None:
            summary.append(
                {
                    "row": row_no,
                    "metric": label,
                    "value": _json_value(c_value),
                }
            )
        values = {
            str(year): _json_value(sheet.Cells(row_no, 8 + offset).Value)
            for offset, year in enumerate(YEARS)
        }
        if any(value is not None for value in values.values()):
            annual.append({"row": row_no, "metric": label, "values": values})
    return {"summary": summary, "annual": annual}


def extract(output: Path = DEFAULT_OUTPUT) -> dict[str, Any]:
    try:
        import win32com.client as win32
    except ImportError as exc:  # pragma: no cover - local Windows dependency
        raise RuntimeError("需要本机 Excel 与 pywin32 才能重算工作簿") from exc

    missing = [str(path) for path in WORKBOOKS.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("缺少铜模型工作簿: " + ", ".join(missing))

    payload: dict[str, Any] = {
        "schema_version": "copper_calculator.workbook_snapshot.v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_contract": (
            "源文件保持不变；仅在临时副本中把以等号开头的中文说明恢复为文本，"
            "再由本机 Excel 全量重算。"
        ),
        "companies": {},
    }
    with tempfile.TemporaryDirectory(prefix="copper_workbook_calc_") as tmp:
        tmp_root = Path(tmp)
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        try:
            for index, (company, source) in enumerate(WORKBOOKS.items()):
                calculation_copy = tmp_root / f"model_{index}.xlsx"
                repairs = _build_excel_readable_copy(source, calculation_copy)
                workbook = excel.Workbooks.Open(
                    str(calculation_copy),
                    UpdateLinks=0,
                    ReadOnly=True,
                    IgnoreReadOnlyRecommended=True,
                )
                try:
                    excel.CalculateFullRebuild()
                    tables = {
                        sheet_name: _annual_table(workbook.Worksheets(sheet_name))
                        for sheet_name in ANNUAL_SHEETS
                    }
                    dcf = _dcf_snapshot(workbook.Worksheets("DCF"))
                finally:
                    workbook.Close(False)
                payload["companies"][company] = {
                    "source_file": source.name,
                    "source_sha256": _sha256(source),
                    "formula_like_prose_repairs": repairs,
                    "tables": tables,
                    "dcf": dcf,
                }
        finally:
            excel.Quit()

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return payload


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    payload = extract(args.output)
    print(
        json.dumps(
            {
                "output": str(args.output),
                "companies": list(payload["companies"]),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
