from __future__ import annotations

"""Compile the lithium workbook into a deployable project-level calculator ledger.

The source workbook is a research reference, not a runtime dependency of the
Viewer deployment.  This compiler preserves each company's mine/salar rows,
stated ownership, 2025-2030 gross LCE volume, full cost and notes in a small JSON
artifact under ``config/``.  Recomputed equity volume deliberately uses the
ownership stated on the same project row.  It therefore also surfaces, rather
than reproduces, formula references in the workbook that point to another
project's ownership cell.
"""

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = ROOT / "config" / "lithium_calculator_project_ledger.json"
YEARS = tuple(range(2025, 2031))
YEAR_COLUMNS = dict(zip(YEARS, range(5, 11), strict=True))

TICKERS = {
    "赣锋锂业": "002460.SZ",
    "天齐锂业": "002466.SZ",
    "中矿资源": "002738.SZ",
    "盛新锂能": "002240.SZ",
    "永兴材料": "002756.SZ",
    "大中矿业": "001203.SZ",
    "国城矿业": "000688.SZ",
    "盐湖股份": "000792.SZ",
    "藏格矿业": "000408.SZ",
    "华友钴业": "603799.SH",
    "雅化集团": "002497.SZ",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _year_values(sheet: Any, row: int) -> dict[str, float]:
    values: dict[str, float] = {}
    for year, column in YEAR_COLUMNS.items():
        value = _number(sheet.cell(row, column).value)
        values[str(year)] = round(value or 0.0, 8)
    return values


def _notes(formula_sheet: Any, start_row: int, end_row: int) -> str:
    notes: list[str] = []
    for row in range(start_row, end_row + 1):
        for column in range(11, min(formula_sheet.max_column, 13) + 1):
            value = formula_sheet.cell(row, column).value
            text = (
                " ".join(str(value).replace("\xa0", " ").split())
                if value is not None else ""
            )
            if text.startswith("="):
                continue
            if text and text not in notes:
                notes.append(text)
    return "；".join(notes)


def _workbook_notes(formula_sheet: Any) -> list[dict[str, Any]]:
    """Preserve every human-written note in workbook columns K-M."""
    notes: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in range(1, formula_sheet.max_row + 1):
        labels: list[str] = []
        for column in range(2, 5):
            value = formula_sheet.cell(row, column).value
            text = (
                " ".join(str(value).replace("\xa0", " ").split())
                if value is not None else ""
            )
            if text and not text.startswith("=") and text not in labels:
                labels.append(text)
        context = " · ".join(labels) or f"第 {row} 行"
        for column in range(11, min(formula_sheet.max_column, 13) + 1):
            cell = formula_sheet.cell(row, column)
            value = cell.value
            text = (
                " ".join(str(value).replace("\xa0", " ").split())
                if value is not None else ""
            )
            if not text or text.startswith("="):
                continue
            identity = (cell.coordinate, text)
            if identity in seen:
                continue
            seen.add(identity)
            notes.append(
                {
                    "cell": cell.coordinate,
                    "row": row,
                    "context": context,
                    "text": text,
                }
            )
    return notes


def _project_rows(formula_sheet: Any) -> tuple[list[int], int]:
    projects: list[int] = []
    total_row = 0
    for row in range(3, formula_sheet.max_row + 1):
        name = str(formula_sheet.cell(row, 2).value or "").strip()
        if "合计" in name:
            total_row = row
            break
        next_label = str(formula_sheet.cell(row + 1, 4).value or "")
        equity_label = str(formula_sheet.cell(row + 2, 4).value or "")
        if (
            name
            and formula_sheet.cell(row, 3).value is not None
            and "碳酸" in next_label
            and "权益" in equity_label
        ):
            projects.append(row)
    if not projects or not total_row:
        raise ValueError(f"{formula_sheet.title}: 未识别到完整锂矿项目区")
    return projects, total_row


def _resource_profit_terms(
    formula_sheet: Any,
    project_rows: list[int],
) -> dict[int, dict[str, Any]]:
    """Read the economic profit attribution embedded in the source workbook.

    The detailed company sheets calculate resource profit from tax-inclusive
    lithium prices and tax-inclusive full costs.  Most projects use the
    company's stated equity volume, while a few rows apply a separate economic
    profit-sharing arrangement.  Corporate income tax is deliberately removed
    from this project-level mapping and applied once after company aggregation
    by the calculator.
    """
    profit_grid_row = next(
        (
            row
            for row in range(1, formula_sheet.max_row + 1)
            if "不同碳酸锂价格对应资源端利润"
            in str(formula_sheet.cell(row, 3).value or "")
        ),
        None,
    )
    if profit_grid_row is None:
        raise ValueError(f"{formula_sheet.title}: 未识别到资源端利润公式")
    formula = str(formula_sheet.cell(profit_grid_row, 5).value or "")
    if not formula.startswith("="):
        raise ValueError(
            f"{formula_sheet.title}!E{profit_grid_row}: 缺少资源端利润公式"
        )

    terms: dict[int, dict[str, Any]] = {}
    for project_row in project_rows:
        cost_row = project_row + 3
        pattern = re.compile(
            rf"\([^)]*-E\${cost_row}/1\.13\)"
            rf"\*E\$(?P<volume_row>\d+)"
            rf"(?P<factors>(?:\*[0-9]+(?:\.[0-9]+)?)*)"
        )
        match = pattern.search(formula)
        if match is None:
            raise ValueError(
                f"{formula_sheet.title}!E{profit_grid_row}: "
                f"未找到成本行 {cost_row} 的利润项"
            )
        volume_row = int(match.group("volume_row"))
        raw_factors = [
            float(value)
            for value in re.findall(
                r"\*([0-9]+(?:\.[0-9]+)?)",
                match.group("factors"),
            )
        ]
        income_tax_factor_present = False
        non_tax_factors: list[float] = []
        for factor in raw_factors:
            if not income_tax_factor_present and abs(factor - 0.85) < 1e-9:
                income_tax_factor_present = True
            else:
                non_tax_factors.append(factor)
        profit_sharing_factor = 1.0
        for factor in non_tax_factors:
            profit_sharing_factor *= factor

        if volume_row == project_row + 1:
            volume_basis = "gross"
        elif volume_row == project_row + 2:
            volume_basis = "equity"
        else:
            raise ValueError(
                f"{formula_sheet.title}!E{profit_grid_row}: "
                f"项目行 {project_row} 使用了无法解释的产量行 {volume_row}"
            )
        terms[project_row] = {
            "profit_volume_basis": volume_basis,
            "profit_sharing_factor": profit_sharing_factor,
            "income_tax_factor_present": income_tax_factor_present,
            "workbook_profit_formula": match.group(0),
        }
    return terms


def _compile_company(formula_sheet: Any, value_sheet: Any) -> dict[str, Any]:
    project_rows, total_row = _project_rows(formula_sheet)
    profit_terms = _resource_profit_terms(formula_sheet, project_rows)
    projects: list[dict[str, Any]] = []
    mismatch_count = 0
    for row in project_rows:
        ownership = _number(value_sheet.cell(row, 3).value)
        if ownership is None:
            raise ValueError(
                f"{formula_sheet.title}!C{row}: 权益公式没有可用缓存值"
            )
        gross = _year_values(value_sheet, row + 1)
        workbook_equity = _year_values(value_sheet, row + 2)
        recomputed_equity = {
            year: round(value * ownership, 8) for year, value in gross.items()
        }
        profit_term = profit_terms[row]
        profit_attribution = (
            profit_term["profit_sharing_factor"]
            if profit_term["profit_volume_basis"] == "gross"
            else ownership * profit_term["profit_sharing_factor"]
        )
        model_equity_start_year = next(
            (
                year
                for year in map(str, YEARS)
                if recomputed_equity[year] > 0
            ),
            None,
        )
        mismatches = {
            year: {
                "workbook": workbook_equity[year],
                "recomputed": recomputed_equity[year],
                "difference": round(
                    recomputed_equity[year] - workbook_equity[year], 8
                ),
                "workbook_formula": str(
                    formula_sheet.cell(row + 2, YEAR_COLUMNS[int(year)]).value
                    or ""
                ),
            }
            for year in map(str, YEARS)
            if abs(recomputed_equity[year] - workbook_equity[year]) > 1e-6
        }
        if mismatches:
            mismatch_count += 1
        projects.append(
            {
                "name": str(formula_sheet.cell(row, 2).value)
                .replace("\n", " ")
                .strip(),
                "type": "resource",
                "enabled": True,
                "ownershipPct": round(ownership * 100.0, 8),
                # Economic profit entitlement is distinct from legal equity.
                # Usually the two are identical.  The source workbook has a
                # small number of explicit profit-sharing adjustments, which
                # must remain visible and editable in the web calculator.
                "profitAttributionPct": round(
                    profit_attribution * 100.0, 8
                ),
                "profitAttributionFollowsOwnership": (
                    profit_term["profit_volume_basis"] == "equity"
                    and abs(
                        profit_term["profit_sharing_factor"] - 1.0
                    ) < 1e-9
                ),
                "profitAttributionBasis": (
                    "项目总产量×单独利润分成"
                    if profit_term["profit_volume_basis"] == "gross"
                    else (
                        "权益产量×项目利润分成"
                        if abs(
                            profit_term["profit_sharing_factor"] - 1.0
                        ) > 1e-9
                        else "权益产量"
                    )
                ),
                "workbookIncomeTaxFactorPresent": profit_term[
                    "income_tax_factor_present"
                ],
                "workbookProfitFormula": profit_term[
                    "workbook_profit_formula"
                ],
                "grossVolumeByYear": gross,
                "costByYear": _year_values(value_sheet, row + 3),
                # Cost-row side notes are often company-level operating notes
                # rather than descriptions of this mine (for example a potash
                # project beside an Argentine salar).  Keep them in
                # workbookNotes, but attach only project/equity-row notes here.
                "note": _notes(formula_sheet, row, row + 2),
                "origin": (
                    f"项目资源底稿：{formula_sheet.title}!"
                    f"B{row}:M{row + 3}"
                ),
                "workbookOwnershipExpression": str(
                    formula_sheet.cell(row, 3).value
                ),
                "workbookEquityVolumeByYear": workbook_equity,
                "recomputedEquityVolumeByYear": recomputed_equity,
                "modelEquityStartYear": (
                    int(model_equity_start_year)
                    if model_equity_start_year is not None else None
                ),
                "ownershipReferenceMismatch": mismatches,
            }
        )

    workbook_total = _year_values(value_sheet, total_row)
    recomputed_total = {
        str(year): round(
            sum(
                project["recomputedEquityVolumeByYear"][str(year)]
                for project in projects
            ),
            8,
        )
        for year in YEARS
    }
    workbook_notes = _workbook_notes(formula_sheet)
    return {
        "name": formula_sheet.title,
        "ticker": TICKERS.get(formula_sheet.title),
        "years": list(YEARS),
        "projects": projects,
        "workbookResourceTotalByYear": workbook_total,
        "recomputedResourceTotalByYear": recomputed_total,
        "projectCount": len(projects),
        "projectWithOwnershipReferenceMismatchCount": mismatch_count,
        "workbookNotes": workbook_notes,
        "workbookNoteCount": len(workbook_notes),
        "workbookTotalRow": total_row,
    }


def compile_ledger(workbook_path: Path) -> dict[str, Any]:
    formula_book = load_workbook(
        workbook_path, data_only=False, read_only=False
    )
    value_book = load_workbook(workbook_path, data_only=True, read_only=False)
    companies: list[dict[str, Any]] = []
    for formula_sheet in formula_book.worksheets:
        if formula_sheet.title not in TICKERS:
            continue
        companies.append(
            _compile_company(
                formula_sheet, value_book[formula_sheet.title]
            )
        )
    missing = sorted(set(TICKERS) - {row["name"] for row in companies})
    if missing:
        raise ValueError(f"工作簿缺少公司页: {', '.join(missing)}")
    return {
        "schema_version": "lithium_calculator.project_ledger.v4",
        "vat_rate_pct": 13.0,
        "income_tax_rate_pct": 15.0,
        "cost_basis": "含税完全成本",
        "resource_profit_formula": (
            "Σ[项目总产量×利润归属比例×"
            "(含税碳酸锂价格/1.13-含税完全成本/1.13)]×(1-15%)"
        ),
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_workbook": workbook_path.name,
        "source_workbook_sha256": _sha256(workbook_path),
        "years": list(YEARS),
        "company_count": len(companies),
        "project_count": sum(row["projectCount"] for row in companies),
        "ownership_reference_mismatch_project_count": sum(
            row["projectWithOwnershipReferenceMismatchCount"]
            for row in companies
        ),
        "companies": companies,
    }


def _find_workbook() -> Path:
    matches = [
        path
        for path in ROOT.glob("*20260606.xlsx")
        if not path.name.startswith("~$")
    ]
    if len(matches) != 1:
        raise FileNotFoundError(
            f"预期唯一的 *20260606.xlsx，实际找到 {len(matches)} 个"
        )
    return matches[0]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    workbook = (args.workbook or _find_workbook()).resolve()
    ledger = compile_ledger(workbook)
    payload = json.dumps(ledger, ensure_ascii=False, indent=2) + "\n"
    if args.check:
        if not args.output.is_file():
            raise SystemExit(f"缺少生成文件: {args.output}")
        if json.loads(args.output.read_text(encoding="utf-8")) != ledger:
            # generated_at_utc is intentionally excluded from semantic checks.
            current = json.loads(args.output.read_text(encoding="utf-8"))
            current.pop("generated_at_utc", None)
            ledger.pop("generated_at_utc", None)
            if current != ledger:
                raise SystemExit("生成文件与当前工作簿不一致")
        return 0
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(payload, encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(args.output),
                "company_count": ledger["company_count"],
                "project_count": ledger["project_count"],
                "ownership_reference_mismatch_project_count": ledger[
                    "ownership_reference_mismatch_project_count"
                ],
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
