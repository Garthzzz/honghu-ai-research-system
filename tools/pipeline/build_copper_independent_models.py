from __future__ import annotations

import argparse
import hashlib
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = (
    ROOT / "cache" / "copper_research" / "models" / "copper_independent_models_v2.json"
)
REFERENCE_WORKBOOK = ROOT / "碳酸锂标的估值测算20260606.xlsx"

FX_USD_CNY = 7.15
LB_PER_TONNE = 2204.62262185
PRICE_GRID_2027_USD_T = (8000.0, 9500.0, 11000.0, 11500.0, 12500.0, 14500.0)


def _canonical_json(payload: Any) -> str:
    return json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _sha256(payload: Any) -> str:
    return "sha256:" + hashlib.sha256(_canonical_json(payload).encode("utf-8")).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return "sha256:" + digest.hexdigest()


def _reference_workbook_contract() -> dict[str, Any]:
    """Validate the user-provided workbook and freeze the transferable formula contract."""
    if not REFERENCE_WORKBOOK.is_file():
        raise FileNotFoundError(f"缺少估值参考工作簿: {REFERENCE_WORKBOOK}")
    workbook = load_workbook(REFERENCE_WORKBOOK, data_only=False, read_only=False)
    expected_sheets = {
        "持仓",
        "Sheet1",
        "汇总",
        "赣锋锂业",
        "天齐锂业",
        "中矿资源",
        "盛新锂能",
        "永兴材料",
        "大中矿业",
        "国城矿业",
        "盐湖股份",
        "藏格矿业",
        "华友钴业",
        "雅化集团",
    }
    missing = sorted(expected_sheets - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"估值参考工作簿缺少工作表: {missing}")
    formula_contract = {
        "project_attributable_output": workbook["赣锋锂业"]["E5"].value,
        "resource_profit_grid": workbook["赣锋锂业"]["E56"].value,
        "other_business_valuation": workbook["赣锋锂业"]["E78"].value,
        "implied_resource_equity_value": workbook["赣锋锂业"]["E83"].value,
        "implied_resource_pe": workbook["赣锋锂业"]["E84"].value,
        "cross_company_resource_value": workbook["汇总"]["D47"].value,
        "cross_company_resource_multiple": workbook["汇总"]["D58"].value,
    }
    if not all(
        isinstance(value, str) and value.startswith("=")
        for value in formula_contract.values()
    ):
        raise ValueError("估值参考工作簿的关键公式合同不完整")
    sheet_summary = []
    for sheet in workbook.worksheets:
        formulas = sum(
            1
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        )
        sheet_summary.append(
            {
                "sheet": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "formula_count": formulas,
            }
        )
    return {
        "path": REFERENCE_WORKBOOK.relative_to(ROOT).as_posix(),
        "sha256": _sha256_file(REFERENCE_WORKBOOK),
        "bytes": REFERENCE_WORKBOOK.stat().st_size,
        "sheet_count": len(workbook.sheetnames),
        "formula_count": sum(row["formula_count"] for row in sheet_summary),
        "sheet_summary": sheet_summary,
        "formula_contract": formula_contract,
        "transfer_policy": (
            "迁移逐项目权益产量、商品价格利润矩阵、其他业务分拆、资源业务隐含估值"
            "和单位权益产量估值；不迁移锂行业13%增值税、碳酸锂折算系数、具体成本、"
            "具体税后系数或锂公司估值倍数。铜模型使用美元/吨同口径价格与成本，"
            "并保留少数股东、资本开支、净债务和多金属残余。"
        ),
    }


def _finite(value: float, name: str) -> float:
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{name} 必须是有限数")
    return number


def _project_margin_usd_mn(
    *,
    production_kt: float,
    ownership: float,
    copper_price_usd_t: float,
    c1_cost_usd_lb: float,
) -> float:
    """项目权益口径铜现金毛利，不等同于会计 EBITDA。"""
    c1_cost_usd_t = c1_cost_usd_lb * LB_PER_TONNE
    return production_kt * ownership * (copper_price_usd_t - c1_cost_usd_t) / 1000.0


def _fcfe_dcf(
    fcfe: list[float],
    *,
    cost_of_equity: float,
    terminal_growth: float,
) -> dict[str, float]:
    if len(fcfe) != 3:
        raise ValueError("FCFE DCF 必须覆盖 FY1—FY3")
    if not 0 <= terminal_growth < cost_of_equity < 1:
        raise ValueError("折现率与永续增长率不满足经济边界")
    discounted = [
        value / ((1.0 + cost_of_equity) ** year)
        for year, value in enumerate(fcfe, start=1)
    ]
    terminal = fcfe[-1] * (1.0 + terminal_growth) / (
        cost_of_equity - terminal_growth
    )
    terminal_pv = terminal / ((1.0 + cost_of_equity) ** 3)
    return {
        "explicit_period_pv": sum(discounted),
        "terminal_value": terminal,
        "terminal_value_pv": terminal_pv,
        "equity_value": sum(discounted) + terminal_pv,
        "terminal_value_share": terminal_pv / (sum(discounted) + terminal_pv),
    }


def _fcfe_dcf_range(
    fcfe: list[float],
    *,
    low_value_cost_of_equity: float,
    low_value_terminal_growth: float,
    high_value_cost_of_equity: float,
    high_value_terminal_growth: float,
) -> dict[str, Any]:
    """Use an explicit discount-rate/growth sensitivity instead of ±x% on value."""
    low = _fcfe_dcf(
        fcfe,
        cost_of_equity=low_value_cost_of_equity,
        terminal_growth=low_value_terminal_growth,
    )
    high = _fcfe_dcf(
        fcfe,
        cost_of_equity=high_value_cost_of_equity,
        terminal_growth=high_value_terminal_growth,
    )
    return {
        "equity_value_low": low["equity_value"],
        "equity_value_high": high["equity_value"],
        "low_value_assumptions": {
            "cost_of_equity": low_value_cost_of_equity,
            "terminal_growth": low_value_terminal_growth,
            "terminal_value_share": low["terminal_value_share"],
        },
        "high_value_assumptions": {
            "cost_of_equity": high_value_cost_of_equity,
            "terminal_growth": high_value_terminal_growth,
            "terminal_value_share": high["terminal_value_share"],
        },
    }


def _justified_pb(*, sustainable_roe: float, cost_of_equity: float, growth: float) -> float:
    if not growth < cost_of_equity:
        raise ValueError("PB—ROE 模型要求永续增长率低于股权成本")
    return (sustainable_roe - growth) / (cost_of_equity - growth)


def _rmb_price_grid(
    *,
    attributable_copper_kt: float,
    cash_cost_usd_t: float,
    baseline_margin_rmb_bn: float,
    actual_net_income_rmb_bn: float,
    other_profit_delta_rmb_bn: float,
    after_tax_conversion: float,
) -> list[dict[str, float]]:
    rows: list[dict[str, float]] = []
    for price in PRICE_GRID_2027_USD_T:
        copper_profit = (
            attributable_copper_kt
            * (price - cash_cost_usd_t)
            / 1000.0
            * FX_USD_CNY
            * after_tax_conversion
            / 1000.0
        )
        total_profit = (
            actual_net_income_rmb_bn
            + copper_profit
            - baseline_margin_rmb_bn
            + other_profit_delta_rmb_bn
        )
        rows.append(
            {
                "copper_price_usd_t": price,
                "copper_after_tax_profit_proxy_rmb_bn": round(copper_profit, 3),
                "non_copper_corporate_residual_rmb_bn": round(
                    total_profit - copper_profit, 3
                ),
                "attributable_net_income_rmb_bn": round(total_profit, 3),
            }
        )
    return rows


def _mmg_price_grid(
    *,
    projects: dict[str, dict[str, float]],
    baseline_margin_usd_bn: float,
    actual_net_income_usd_bn: float,
    other_profit_delta_usd_bn: float,
    after_tax_conversion: float,
) -> list[dict[str, float]]:
    rows: list[dict[str, float]] = []
    for price in PRICE_GRID_2027_USD_T:
        copper_cash_margin = sum(
            _project_margin_usd_mn(
                production_kt=values["production_kt"],
                ownership=values["ownership"],
                copper_price_usd_t=price,
                c1_cost_usd_lb=values["c1_usd_lb"],
            )
            for values in projects.values()
        ) / 1000.0
        copper_after_tax = copper_cash_margin * after_tax_conversion
        total_profit = (
            actual_net_income_usd_bn
            + (copper_cash_margin - baseline_margin_usd_bn) * after_tax_conversion
            + other_profit_delta_usd_bn
        )
        rows.append(
            {
                "copper_price_usd_t": price,
                "copper_equity_cash_margin_usd_bn": round(copper_cash_margin, 4),
                "copper_after_tax_profit_proxy_usd_bn": round(copper_after_tax, 4),
                "non_copper_corporate_finance_residual_usd_bn": round(
                    total_profit - copper_after_tax, 4
                ),
                "attributable_net_income_usd_bn": round(total_profit, 4),
            }
        )
    return rows


PRICE_SCENARIOS = {
    "下行情景": {2026: 9500.0, 2027: 9000.0, 2028: 9000.0},
    "基准情景": {2026: 12500.0, 2027: 11500.0, 2028: 11000.0},
    "上行情景": {2026: 14500.0, 2027: 14500.0, 2028: 14000.0},
}


def _zijin_model() -> dict[str, Any]:
    actual = {
        "year": 2025,
        "revenue_rmb_bn": 349.079,
        "attributable_net_income_rmb_bn": 51.777,
        "operating_cash_flow_rmb_bn": 75.430,
        "capex_cash_rmb_bn": 30.982,
        "parent_equity_rmb_bn": 185.542,
        "total_assets_rmb_bn": 512.005,
        "mine_copper_kt_100pct": 1085.126,
        "attributable_copper_kt": 885.569,
    }
    source_ref = (
        "papers/铜/2026-04-27_紫金矿业_2025年年度报告_官方英文.pdf"
    )
    q1_source_ref = (
        "papers/铜/2026-04-27_紫金矿业_2026年第一季度报告_官方英文.pdf"
    )
    h1_source_ref = "https://www.zijinmining.com/news/news-detail-122842.htm"

    # 权益产量显式处理卡莫阿、塞尔维亚、科卢韦齐、巨龙及其他矿山。
    attributable_copper = {
        2026: {
            "kamoa": 310.0 * 0.442,
            "serbia": 296.0 * 0.8457,
            "kolwezi": 109.0 * 0.67,
            "julong": 300.0 * 0.5816,
            "other": 280.0,
        },
        2027: {
            "kamoa": 370.0 * 0.442,
            "serbia": 315.0 * 0.8457,
            "kolwezi": 112.0 * 0.67,
            "julong": 330.0 * 0.5816,
            "other": 295.0,
        },
        2028: {
            "kamoa": 410.0 * 0.442,
            "serbia": 345.0 * 0.8457,
            "kolwezi": 115.0 * 0.67,
            "julong": 350.0 * 0.5816,
            "zhunuo_xiongcun_ramp": 70.0,
            "other": 305.0,
        },
    }
    attributed = {year: sum(parts.values()) for year, parts in attributable_copper.items()}
    baseline_price = 9950.0
    baseline_cash_cost = 4400.0
    after_tax_conversion = 0.62
    baseline_copper_margin = (
        actual["attributable_copper_kt"]
        * (baseline_price - baseline_cash_cost)
        / 1000.0
        * FX_USD_CNY
        * after_tax_conversion
        / 1000.0
    )
    non_copper_delta = {
        2026: 17.5,
        2027: 27.0,
        2028: 35.0,
    }
    revenue = {2026: 430.0, 2027: 475.0, 2028: 535.0}
    fcfe = {2026: 62.0, 2027: 69.0, 2028: 80.0}
    capex = {2026: 45.0, 2027: 50.0, 2028: 55.0}
    parent_equity = {2026: 237.0, 2027: 294.0, 2028: 358.0}
    total_assets = {2026: 565.0, 2027: 625.0, 2028: 690.0}
    cash_cost = {2026: 4600.0, 2027: 4500.0, 2028: 4400.0}

    scenarios: dict[str, list[dict[str, Any]]] = {}
    for scenario, prices in PRICE_SCENARIOS.items():
        rows: list[dict[str, Any]] = []
        for year in (2026, 2027, 2028):
            copper_margin = (
                attributed[year]
                * (prices[year] - cash_cost[year])
                / 1000.0
                * FX_USD_CNY
                * after_tax_conversion
                / 1000.0
            )
            if scenario == "下行情景":
                other = non_copper_delta[year] * 0.62
            elif scenario == "上行情景":
                other = non_copper_delta[year] * 1.28
            else:
                other = non_copper_delta[year]
            net_income = (
                actual["attributable_net_income_rmb_bn"]
                + copper_margin
                - baseline_copper_margin
                + other
            )
            average_equity = (
                (actual["parent_equity_rmb_bn"] + parent_equity[year]) / 2.0
                if year == 2026
                else (parent_equity[year - 1] + parent_equity[year]) / 2.0
            )
            average_assets = (
                (actual["total_assets_rmb_bn"] + total_assets[year]) / 2.0
                if year == 2026
                else (total_assets[year - 1] + total_assets[year]) / 2.0
            )
            rows.append(
                {
                    "year": year,
                    "copper_price_usd_t": prices[year],
                    "attributable_copper_kt": round(attributed[year], 1),
                    "copper_cash_cost_usd_t": cash_cost[year],
                    "copper_after_tax_contribution_rmb_bn": round(copper_margin, 2),
                    "non_copper_profit_delta_rmb_bn": round(other, 2),
                    "revenue_rmb_bn": round(
                        revenue[year]
                        * (
                            0.89
                            if scenario == "下行情景"
                            else 1.10 if scenario == "上行情景" else 1.0
                        ),
                        2,
                    ),
                    "attributable_net_income_rmb_bn": round(net_income, 2),
                    "fcfe_rmb_bn": round(
                        fcfe[year]
                        * (
                            0.72
                            if scenario == "下行情景"
                            else 1.24 if scenario == "上行情景" else 1.0
                        ),
                        2,
                    ),
                    "capex_rmb_bn": capex[year],
                    "average_parent_equity_rmb_bn": round(average_equity, 2),
                    "ending_parent_equity_rmb_bn": parent_equity[year],
                    "average_total_assets_rmb_bn": round(average_assets, 2),
                    "roe": round(net_income / average_equity, 4),
                    "roa": round(net_income / average_assets, 4),
                }
            )
        scenarios[scenario] = rows

    base_rows = scenarios["基准情景"]
    normalized_income = base_rows[1]["attributable_net_income_rmb_bn"]
    fcfe_series = [row["fcfe_rmb_bn"] for row in base_rows]
    dcf = _fcfe_dcf(
        fcfe_series,
        cost_of_equity=0.105,
        terminal_growth=0.03,
    )
    dcf_range = _fcfe_dcf_range(
        fcfe_series,
        low_value_cost_of_equity=0.115,
        low_value_terminal_growth=0.025,
        high_value_cost_of_equity=0.095,
        high_value_terminal_growth=0.03,
    )
    pb_roe_low = 0.22
    pb_roe_high = 0.28
    pb_low = _justified_pb(
        sustainable_roe=pb_roe_low, cost_of_equity=0.105, growth=0.03
    )
    pb_high = _justified_pb(
        sustainable_roe=pb_roe_high, cost_of_equity=0.105, growth=0.03
    )
    adopted_pb_low = round(pb_low, 2)
    adopted_pb_high = round(pb_high, 2)
    price_grid = _rmb_price_grid(
        attributable_copper_kt=attributed[2027],
        cash_cost_usd_t=cash_cost[2027],
        baseline_margin_rmb_bn=baseline_copper_margin,
        actual_net_income_rmb_bn=actual["attributable_net_income_rmb_bn"],
        other_profit_delta_rmb_bn=non_copper_delta[2027],
        after_tax_conversion=after_tax_conversion,
    )
    return {
        "company": "紫金矿业",
        "ticker": "601899.SH",
        "currency": "CNY",
        "actual_2025": actual,
        "sources": [source_ref, q1_source_ref, h1_source_ref],
        "model_method": (
            "归母净利润＝2025归母净利润＋权益铜现金毛利变化×税后转化率"
            "＋黄金、锂及其他金属增量；权益产量按项目权益逐项计算。"
            "参照商品资源股工作簿，另做2027年铜价—资源利润矩阵；当前市值和"
            "其他业务估值只在冻结后的外部对账层进入。"
        ),
        "critical_inputs": {
            "price_scenarios_usd_t": PRICE_SCENARIOS,
            "attributable_copper_breakdown_kt": attributable_copper,
            "copper_cash_cost_usd_t": cash_cost,
            "baseline_copper_price_usd_t": baseline_price,
            "baseline_copper_cash_cost_usd_t": baseline_cash_cost,
            "after_tax_conversion": after_tax_conversion,
            "non_copper_profit_delta_rmb_bn": non_copper_delta,
            "non_copper_basis": (
                "以2026年105吨矿产金、120万吨矿产铜、12万吨LCE等官方计划，"
                "以及2026H1归母净利润39.1亿元、矿产铜53.4万吨的官方预告校验；"
                "不把锂与金的增量伪装成铜盈利。"
            ),
        },
        "scenarios": scenarios,
        "valuation": {
            "normalized_year": 2027,
            "normalized_net_income_rmb_bn": normalized_income,
            "normalized_pe": {
                "multiple_low": 10.0,
                "multiple_high": 14.0,
                "equity_value_low_rmb_bn": round(normalized_income * 10.0, 2),
                "equity_value_high_rmb_bn": round(normalized_income * 14.0, 2),
                "parameter_basis": (
                    "2027正常化利润对应的周期中段倍数。下限高于2026-07-24 Wind "
                    "前瞻PE 9.14倍，用于要求项目兑现而非沿用即期折价；上限以中金"
                    "2026-07-15目标价所用2026E 14.2倍PE为外部上沿参照，取整为14倍。"
                ),
            },
            "pb_roe": {
                "sustainable_roe_low": pb_roe_low,
                "sustainable_roe_high": pb_roe_high,
                "cost_of_equity": 0.105,
                "terminal_growth": 0.03,
                "formula_justified_pb_low": adopted_pb_low,
                "formula_justified_pb_high": adopted_pb_high,
                "adopted_pb_low": adopted_pb_low,
                "adopted_pb_high": adopted_pb_high,
                "forward_parent_equity_rmb_bn": parent_equity[2027],
                "equity_value_low_rmb_bn": round(
                    parent_equity[2027] * adopted_pb_low, 2
                ),
                "equity_value_high_rmb_bn": round(
                    parent_equity[2027] * adopted_pb_high, 2
                ),
                "parameter_basis": (
                    "2027模型ROE为高景气水平，不直接进入永续期。可持续ROE下限22%"
                    "对应铜金价格和扩产回报正常化，上限28%仍低于2026-07-24 Wind "
                    "TTM ROE 30.78%及模型2027 ROE；股权成本10.5%、长期增长3.0%。"
                ),
            },
            "fcfe_dcf": {
                **{key: round(value, 4) for key, value in dcf.items()},
                "equity_value_low": round(dcf_range["equity_value_low"], 4),
                "equity_value_high": round(dcf_range["equity_value_high"], 4),
                "low_value_assumptions": dcf_range["low_value_assumptions"],
                "high_value_assumptions": dcf_range["high_value_assumptions"],
                "unit": "RMB bn",
                "cost_of_equity": 0.105,
                "terminal_growth": 0.03,
                "parameter_basis": (
                    "下限使用股权成本11.5%、永续增长2.5%；上限使用股权成本9.5%、"
                    "永续增长3.0%。区间来自折现率与终值增长敏感性，不再对DCF结果"
                    "机械加减15%。Morgan Stanley 2026-07-10 DCF的WACC 7.5%与"
                    "长期收入增长3.0%只作外部对照；本模型使用股权自由现金流，故用"
                    "高于WACC的股权成本。"
                ),
            },
            "workbook_style_commodity_bridge": {
                "role": "resource_earnings_and_implied_valuation_input",
                "normalized_year": 2027,
                "attributable_copper_kt": round(attributed[2027], 1),
                "portfolio_cash_cost_usd_t": cash_cost[2027],
                "price_sensitivity": price_grid,
                "sotp_readiness": "simplified_residual_business_range",
                "formula": (
                    "铜资源税后利润代理＝权益铜产量×（铜价－组合现金成本）"
                    "×美元兑人民币×税后转化率；其他业务与公司层残余＝"
                    "归母净利润－铜资源税后利润代理。"
                ),
                "limitation": (
                    "紫金未披露所有矿山同口径完全成本，使用组合现金成本；"
                    "黄金、锂、锌铅银和公司层费用只能合并为残余，不构成逐分部NAV。"
                ),
            },
        },
        "limitations": [
            "矿山成本披露口径不完全一致，模型用现金毛利变化桥接归母利润，不替代完整矿山NAV。",
            "卡莫阿恢复节奏、巨龙爬坡和金锂价格会使2026—2028年利润偏离基准。",
            "H1数据为公司业绩预告，正式中报发布后应建立修订而不是覆盖原冻结版本。",
        ],
    }


def _cmoc_model() -> dict[str, Any]:
    actual = {
        "year": 2025,
        "revenue_rmb_bn": 206.684,
        "attributable_net_income_rmb_bn": 20.339,
        "operating_cash_flow_rmb_bn": 20.843,
        "parent_equity_rmb_bn": 82.435,
        "total_assets_rmb_bn": 200.932,
        "copper_kt_100pct": 741.149,
    }
    source_ref = "papers/铜/2026-03-28_洛阳钼业_2025年年度报告_官方.pdf"
    q1_source_ref = "papers/铜/2026-04-25_洛阳钼业_2026年第一季度报告_官方.pdf"
    ownership_factor = 0.76
    production_100pct = {2026: 790.0, 2027: 890.0, 2028: 960.0}
    attributable = {
        year: value * ownership_factor for year, value in production_100pct.items()
    }
    baseline_attributable = actual["copper_kt_100pct"] * ownership_factor
    baseline_price = 9950.0
    baseline_cash_cost = 4660.0
    cash_cost = {2026: 4750.0, 2027: 4500.0, 2028: 4400.0}
    after_tax_conversion = 0.60
    baseline_margin = (
        baseline_attributable
        * (baseline_price - baseline_cash_cost)
        / 1000.0
        * FX_USD_CNY
        * after_tax_conversion
        / 1000.0
    )
    non_copper_delta = {2026: 5.0, 2027: 8.0, 2028: 10.0}
    revenue = {2026: 255.0, 2027: 270.0, 2028: 295.0}
    fcfe = {2026: 22.0, 2027: 25.0, 2028: 27.0}
    capex = {2026: 16.0, 2027: 18.0, 2028: 20.0}
    parent_equity = {2026: 104.5, 2027: 129.0, 2028: 155.0}
    total_assets = {2026: 225.0, 2027: 245.0, 2028: 270.0}

    scenarios: dict[str, list[dict[str, Any]]] = {}
    for scenario, prices in PRICE_SCENARIOS.items():
        rows: list[dict[str, Any]] = []
        for year in (2026, 2027, 2028):
            margin = (
                attributable[year]
                * (prices[year] - cash_cost[year])
                / 1000.0
                * FX_USD_CNY
                * after_tax_conversion
                / 1000.0
            )
            other = non_copper_delta[year] * (
                0.45 if scenario == "下行情景" else 1.30 if scenario == "上行情景" else 1.0
            )
            net_income = actual["attributable_net_income_rmb_bn"] + margin - baseline_margin + other
            average_equity = (
                (actual["parent_equity_rmb_bn"] + parent_equity[year]) / 2.0
                if year == 2026
                else (parent_equity[year - 1] + parent_equity[year]) / 2.0
            )
            average_assets = (
                (actual["total_assets_rmb_bn"] + total_assets[year]) / 2.0
                if year == 2026
                else (total_assets[year - 1] + total_assets[year]) / 2.0
            )
            rows.append(
                {
                    "year": year,
                    "copper_price_usd_t": prices[year],
                    "copper_production_kt_100pct": production_100pct[year],
                    "attributable_copper_kt_proxy": round(attributable[year], 1),
                    "copper_cash_cost_usd_t": cash_cost[year],
                    "copper_after_tax_contribution_rmb_bn": round(margin, 2),
                    "non_copper_profit_delta_rmb_bn": round(other, 2),
                    "revenue_rmb_bn": round(
                        revenue[year]
                        * (
                            0.90
                            if scenario == "下行情景"
                            else 1.09 if scenario == "上行情景" else 1.0
                        ),
                        2,
                    ),
                    "attributable_net_income_rmb_bn": round(net_income, 2),
                    "fcfe_rmb_bn": round(
                        fcfe[year]
                        * (
                            0.64
                            if scenario == "下行情景"
                            else 1.25 if scenario == "上行情景" else 1.0
                        ),
                        2,
                    ),
                    "capex_rmb_bn": capex[year],
                    "average_parent_equity_rmb_bn": round(average_equity, 2),
                    "ending_parent_equity_rmb_bn": parent_equity[year],
                    "average_total_assets_rmb_bn": round(average_assets, 2),
                    "roe": round(net_income / average_equity, 4),
                    "roa": round(net_income / average_assets, 4),
                }
            )
        scenarios[scenario] = rows

    base_rows = scenarios["基准情景"]
    normalized_income = base_rows[1]["attributable_net_income_rmb_bn"]
    fcfe_series = [row["fcfe_rmb_bn"] for row in base_rows]
    dcf = _fcfe_dcf(
        fcfe_series,
        cost_of_equity=0.11,
        terminal_growth=0.025,
    )
    dcf_range = _fcfe_dcf_range(
        fcfe_series,
        low_value_cost_of_equity=0.12,
        low_value_terminal_growth=0.02,
        high_value_cost_of_equity=0.10,
        high_value_terminal_growth=0.025,
    )
    pb_roe_low = 0.19
    pb_roe_high = 0.25
    pb_low = _justified_pb(
        sustainable_roe=pb_roe_low, cost_of_equity=0.11, growth=0.025
    )
    pb_high = _justified_pb(
        sustainable_roe=pb_roe_high, cost_of_equity=0.11, growth=0.025
    )
    adopted_pb_low = round(pb_low, 2)
    adopted_pb_high = round(pb_high, 2)
    price_grid = _rmb_price_grid(
        attributable_copper_kt=attributable[2027],
        cash_cost_usd_t=cash_cost[2027],
        baseline_margin_rmb_bn=baseline_margin,
        actual_net_income_rmb_bn=actual["attributable_net_income_rmb_bn"],
        other_profit_delta_rmb_bn=non_copper_delta[2027],
        after_tax_conversion=after_tax_conversion,
    )
    return {
        "company": "洛阳钼业",
        "ticker": "603993.SH",
        "currency": "CNY",
        "actual_2025": actual,
        "sources": [source_ref, q1_source_ref],
        "model_method": (
            "归母净利润＝2025归母净利润＋TFM/KFM权益铜现金毛利变化×税后转化率"
            "＋钴、金、钼钨铌磷和IXM增量；贸易收入单列为低利润率业务，不按矿山毛利率估值。"
            "参照商品资源股工作簿，另做2027年铜价—资源利润矩阵；当前市值和"
            "其他业务估值只在冻结后的外部对账层进入。"
        ),
        "critical_inputs": {
            "price_scenarios_usd_t": PRICE_SCENARIOS,
            "copper_production_kt_100pct": production_100pct,
            "ownership_factor_proxy": ownership_factor,
            "ownership_factor_basis": "TFM权益80%、KFM权益71.25%，按两矿组合近似76%。",
            "copper_cash_cost_usd_t": cash_cost,
            "after_tax_conversion": after_tax_conversion,
            "non_copper_profit_delta_rmb_bn": non_copper_delta,
            "non_copper_basis": (
                "2026年官方铜产量指引76—82万吨；KFM二期预计2027年建成并新增年均"
                "10万吨铜。2026Q1铜产量18.79万吨、归母净利润77.60亿元用于校验，"
                "未将单季高价和贸易营运资金释放机械年化。"
            ),
        },
        "scenarios": scenarios,
        "valuation": {
            "normalized_year": 2027,
            "normalized_net_income_rmb_bn": normalized_income,
            "normalized_pe": {
                "multiple_low": 9.0,
                "multiple_high": 12.0,
                "equity_value_low_rmb_bn": round(normalized_income * 9.0, 2),
                "equity_value_high_rmb_bn": round(normalized_income * 12.0, 2),
                "parameter_basis": (
                    "近期机构在2026-07-10至07-12的2027E市场PE约8.0—9.7倍，"
                    "Wind 2026-07-24前瞻PE为11.45倍。下限取9倍覆盖周期折价，"
                    "上限取12倍覆盖KFM二期兑现但不把当前高铜钴ROE永久化。"
                ),
            },
            "pb_roe": {
                "sustainable_roe_low": pb_roe_low,
                "sustainable_roe_high": pb_roe_high,
                "cost_of_equity": 0.11,
                "terminal_growth": 0.025,
                "formula_justified_pb_low": adopted_pb_low,
                "formula_justified_pb_high": adopted_pb_high,
                "adopted_pb_low": adopted_pb_low,
                "adopted_pb_high": adopted_pb_high,
                "forward_parent_equity_rmb_bn": parent_equity[2027],
                "equity_value_low_rmb_bn": round(
                    parent_equity[2027] * adopted_pb_low, 2
                ),
                "equity_value_high_rmb_bn": round(
                    parent_equity[2027] * adopted_pb_high, 2
                ),
                "parameter_basis": (
                    "可持续ROE取19%—25%，低于2026-07-24 Wind TTM ROE 27.09%"
                    "及模型扩产期ROE，以反映刚果（金）税费、电力、硫酸和现金汇回"
                    "风险；股权成本11.0%、长期增长2.5%。"
                ),
            },
            "fcfe_dcf": {
                **{key: round(value, 4) for key, value in dcf.items()},
                "equity_value_low": round(dcf_range["equity_value_low"], 4),
                "equity_value_high": round(dcf_range["equity_value_high"], 4),
                "low_value_assumptions": dcf_range["low_value_assumptions"],
                "high_value_assumptions": dcf_range["high_value_assumptions"],
                "unit": "RMB bn",
                "cost_of_equity": 0.11,
                "terminal_growth": 0.025,
                "parameter_basis": (
                    "下限使用股权成本12.0%、永续增长2.0%；上限使用股权成本10.0%、"
                    "永续增长2.5%。BofA 2026-07-10 DCF采用WACC 8.3%、终值增长1%，"
                    "本模型因折现的是股权现金流而采用更高股权成本。"
                ),
            },
            "workbook_style_commodity_bridge": {
                "role": "resource_earnings_and_implied_valuation_input",
                "normalized_year": 2027,
                "attributable_copper_kt_proxy": round(attributable[2027], 1),
                "portfolio_cash_cost_usd_t": cash_cost[2027],
                "price_sensitivity": price_grid,
                "sotp_readiness": "simplified_residual_business_range",
                "formula": (
                    "铜资源税后利润代理＝TFM/KFM权益铜代理×（铜价－组合现金成本）"
                    "×美元兑人民币×税后转化率；其他业务与公司层残余＝"
                    "归母净利润－铜资源税后利润代理。"
                ),
                "limitation": (
                    "公开资料没有按TFM/KFM逐年给出完全可比产量与成本，"
                    "因此保留76%组合权益代理；钴副产品抵扣、IXM、钼钨铌磷及"
                    "公司层费用合并为残余，不能解释成纯粹的非铜利润。"
                ),
            },
        },
        "limitations": [
            "TFM与KFM未按所有期间逐矿披露完全可比成本，76%权益系组合近似。",
            "钴出口配额、硫酸和电力变化可能使铜钴联合成本与营运资金显著偏离。",
            "IXM营业收入大但利润率低，模型不以总收入增速代替矿山利润增长。",
        ],
    }


def _mmg_model() -> dict[str, Any]:
    actual = {
        "year": 2025,
        "revenue_usd_bn": 6.218,
        "attributable_net_income_usd_bn": 0.5094,
        "operating_cash_flow_usd_bn": 2.6895,
        "capex_cash_usd_bn": 1.0796,
        "parent_equity_usd_bn": 3.9599,
        "total_equity_usd_bn": 6.8999,
        "total_assets_usd_bn": 15.3005,
    }
    source_ref = "papers/铜/2026-04-21_五矿资源_2025年年度报告_官方英文.pdf"
    q2_source_ref = "papers/铜/2026-07-21_五矿资源_2026年第二季度生产报告_官方英文.pdf"
    projects = {
        2025: {
            "Las Bambas": {"production_kt": 410.834, "ownership": 0.625, "c1_usd_lb": 1.12},
            "Kinsevere": {"production_kt": 52.791, "ownership": 1.0, "c1_usd_lb": 3.12},
            "Khoemacau": {"production_kt": 42.120, "ownership": 0.55, "c1_usd_lb": 1.97},
        },
        2026: {
            "Las Bambas": {"production_kt": 390.0, "ownership": 0.625, "c1_usd_lb": 0.95},
            "Kinsevere": {"production_kt": 70.0, "ownership": 1.0, "c1_usd_lb": 2.70},
            "Khoemacau": {"production_kt": 50.5, "ownership": 0.55, "c1_usd_lb": 1.85},
        },
        2027: {
            "Las Bambas": {"production_kt": 400.0, "ownership": 0.625, "c1_usd_lb": 1.05},
            "Kinsevere": {"production_kt": 72.0, "ownership": 1.0, "c1_usd_lb": 2.55},
            "Khoemacau": {"production_kt": 65.0, "ownership": 0.55, "c1_usd_lb": 1.75},
        },
        2028: {
            "Las Bambas": {"production_kt": 410.0, "ownership": 0.625, "c1_usd_lb": 1.10},
            "Kinsevere": {"production_kt": 75.0, "ownership": 1.0, "c1_usd_lb": 2.45},
            "Khoemacau": {"production_kt": 105.0, "ownership": 0.55, "c1_usd_lb": 1.60},
        },
    }
    baseline_price = 9950.0
    baseline_margin = sum(
        _project_margin_usd_mn(
            production_kt=values["production_kt"],
            ownership=values["ownership"],
            copper_price_usd_t=baseline_price,
            c1_cost_usd_lb=values["c1_usd_lb"],
        )
        for values in projects[2025].values()
    ) / 1000.0
    after_tax_conversion = 0.47
    non_copper_delta = {2026: 0.10, 2027: 0.06, 2028: 0.14}
    revenue = {2026: 7.80, 2027: 7.30, 2028: 7.75}
    fcfe = {2026: 1.20, 2027: 1.00, 2028: 1.10}
    capex = {2026: 1.25, 2027: 1.45, 2028: 1.30}
    parent_equity = {2026: 5.05, 2027: 5.96, 2028: 6.94}
    total_assets = {2026: 16.2, 2027: 17.0, 2028: 17.8}

    scenarios: dict[str, list[dict[str, Any]]] = {}
    for scenario, prices in PRICE_SCENARIOS.items():
        rows: list[dict[str, Any]] = []
        for year in (2026, 2027, 2028):
            project_margin = sum(
                _project_margin_usd_mn(
                    production_kt=values["production_kt"],
                    ownership=values["ownership"],
                    copper_price_usd_t=prices[year],
                    c1_cost_usd_lb=values["c1_usd_lb"],
                )
                for values in projects[year].values()
            ) / 1000.0
            other = non_copper_delta[year] * (
                0.50 if scenario == "下行情景" else 1.35 if scenario == "上行情景" else 1.0
            )
            net_income = (
                actual["attributable_net_income_usd_bn"]
                + (project_margin - baseline_margin) * after_tax_conversion
                + other
            )
            average_equity = (
                (actual["parent_equity_usd_bn"] + parent_equity[year]) / 2.0
                if year == 2026
                else (parent_equity[year - 1] + parent_equity[year]) / 2.0
            )
            average_assets = (
                (actual["total_assets_usd_bn"] + total_assets[year]) / 2.0
                if year == 2026
                else (total_assets[year - 1] + total_assets[year]) / 2.0
            )
            rows.append(
                {
                    "year": year,
                    "copper_price_usd_t": prices[year],
                    "project_inputs": projects[year],
                    "attributable_copper_kt": round(
                        sum(
                            values["production_kt"] * values["ownership"]
                            for values in projects[year].values()
                        ),
                        1,
                    ),
                    "copper_equity_cash_margin_usd_bn": round(project_margin, 3),
                    "non_copper_profit_delta_usd_bn": round(other, 3),
                    "revenue_usd_bn": round(
                        revenue[year]
                        * (
                            0.88
                            if scenario == "下行情景"
                            else 1.11 if scenario == "上行情景" else 1.0
                        ),
                        3,
                    ),
                    "attributable_net_income_usd_bn": round(net_income, 3),
                    "fcfe_usd_bn": round(
                        fcfe[year]
                        * (
                            0.52
                            if scenario == "下行情景"
                            else 1.30 if scenario == "上行情景" else 1.0
                        ),
                        3,
                    ),
                    "capex_usd_bn": capex[year],
                    "average_parent_equity_usd_bn": round(average_equity, 3),
                    "ending_parent_equity_usd_bn": parent_equity[year],
                    "average_total_assets_usd_bn": round(average_assets, 3),
                    "roe": round(net_income / average_equity, 4),
                    "roa": round(net_income / average_assets, 4),
                }
            )
        scenarios[scenario] = rows

    base_rows = scenarios["基准情景"]
    normalized_income = base_rows[1]["attributable_net_income_usd_bn"]
    fcfe_series = [row["fcfe_usd_bn"] for row in base_rows]
    dcf = _fcfe_dcf(
        fcfe_series,
        cost_of_equity=0.13,
        terminal_growth=0.02,
    )
    dcf_range = _fcfe_dcf_range(
        fcfe_series,
        low_value_cost_of_equity=0.145,
        low_value_terminal_growth=0.015,
        high_value_cost_of_equity=0.115,
        high_value_terminal_growth=0.02,
    )
    pb_roe_low = 0.11
    pb_roe_high = 0.16
    pb_low = _justified_pb(
        sustainable_roe=pb_roe_low, cost_of_equity=0.13, growth=0.02
    )
    pb_high = _justified_pb(
        sustainable_roe=pb_roe_high, cost_of_equity=0.13, growth=0.02
    )
    adopted_pb_low = round(pb_low, 2)
    adopted_pb_high = round(pb_high, 2)
    price_grid = _mmg_price_grid(
        projects=projects[2027],
        baseline_margin_usd_bn=baseline_margin,
        actual_net_income_usd_bn=actual["attributable_net_income_usd_bn"],
        other_profit_delta_usd_bn=non_copper_delta[2027],
        after_tax_conversion=after_tax_conversion,
    )
    return {
        "company": "五矿资源",
        "ticker": "1208.HK",
        "currency": "USD",
        "actual_2025": actual,
        "sources": [source_ref, q2_source_ref],
        "model_method": (
            "逐矿权益铜现金毛利＝项目产量×权益比例×（铜价－C1成本）；"
            "归母净利润＝2025归母净利润＋权益铜现金毛利变化×税后转化率"
            "＋锌、金、银及其他业务增量。另按参考工作簿结构生成2027年铜价—"
            "资源利润矩阵；当前市值和隐含估值只在独立模型冻结后的对账层使用。"
        ),
        "critical_inputs": {
            "price_scenarios_usd_t": PRICE_SCENARIOS,
            "project_production_ownership_and_c1": projects,
            "after_tax_conversion": after_tax_conversion,
            "non_copper_profit_delta_usd_bn": non_copper_delta,
            "non_copper_basis": (
                "Las Bambas、Kinsevere、Khoemacau分别按62.5%、100%、55%权益；"
                "2026年产量和C1成本采用公司2026Q2更新后的指引中值。"
            ),
        },
        "scenarios": scenarios,
        "valuation": {
            "normalized_year": 2027,
            "normalized_net_income_usd_bn": normalized_income,
            "normalized_pe": {
                "multiple_low": 8.0,
                "multiple_high": 11.0,
                "equity_value_low_usd_bn": round(normalized_income * 8.0, 3),
                "equity_value_high_usd_bn": round(normalized_income * 11.0, 3),
                "parameter_basis": (
                    "Morgan Stanley、Citi与Jefferies在2026-07-21至07-22披露的"
                    "2027E PE约8.3—9.6倍；独立利润低于卖方路径，因此上限只扩至"
                    "11倍以容纳Khoemacau爬坡和去杠杆，不使用更高周期溢价。"
                ),
            },
            "pb_roe": {
                "sustainable_roe_low": pb_roe_low,
                "sustainable_roe_high": pb_roe_high,
                "cost_of_equity": 0.13,
                "terminal_growth": 0.02,
                "formula_justified_pb_low": adopted_pb_low,
                "formula_justified_pb_high": adopted_pb_high,
                "adopted_pb_low": adopted_pb_low,
                "adopted_pb_high": adopted_pb_high,
                "forward_parent_equity_usd_bn": parent_equity[2027],
                "equity_value_low_usd_bn": round(
                    parent_equity[2027] * adopted_pb_low, 3
                ),
                "equity_value_high_usd_bn": round(
                    parent_equity[2027] * adopted_pb_high, 3
                ),
                "parameter_basis": (
                    "可持续ROE取11%—16%，低于模型扩产期回报；股权成本13%、长期"
                    "增长2%。区间显式反映高净债务、少数股东、秘鲁社区风险和"
                    "刚果（金）成本，不再把高于公式单点的PB写成周期折价。"
                ),
            },
            "fcfe_dcf": {
                **{key: round(value, 4) for key, value in dcf.items()},
                "equity_value_low": round(dcf_range["equity_value_low"], 4),
                "equity_value_high": round(dcf_range["equity_value_high"], 4),
                "low_value_assumptions": dcf_range["low_value_assumptions"],
                "high_value_assumptions": dcf_range["high_value_assumptions"],
                "unit": "USD bn",
                "cost_of_equity": 0.13,
                "terminal_growth": 0.02,
                "parameter_basis": (
                    "下限使用股权成本14.5%、永续增长1.5%；上限使用股权成本11.5%、"
                    "永续增长2.0%。Morgan Stanley 2026-07-21使用股权成本16.9%、"
                    "长期收入增长2%，Citi使用WACC 9.5%；本区间位于两者风险口径之间。"
                ),
            },
            "workbook_style_commodity_bridge": {
                "role": "独立资源利润敏感性；市场隐含估值在冻结后计算",
                "year": 2027,
                "attributable_copper_kt": round(
                    sum(
                        values["production_kt"] * values["ownership"]
                        for values in projects[2027].values()
                    ),
                    1,
                ),
                "price_sensitivity": price_grid,
                "sotp_readiness": "资源主导集团，仅允许集团盈利口径诊断",
                "formula": (
                    "权益铜现金毛利＝Σ[项目产量×权益比例×（铜价－C1成本）]；"
                    "权益铜税后利润代理＝权益铜现金毛利×税后转化率。"
                ),
                "limitation": (
                    "五矿资源的少数股东、利息、总部费用和其他金属共同形成负的"
                    "公司层残余，不能照搬锂工作簿把它资本化成正的“其他业务价值”；"
                    "冻结后只用集团归母盈利、当前市值和单位权益铜产量作诊断。"
                ),
            },
        },
        "limitations": [
            "C1成本已扣伴生品，银金价格变化会改变成本而非只改变收入。",
            "Khoemacau扩建首批精矿计划在2028H1，2028模型只按部分爬坡，不按13万吨满产。",
            "Las Bambas少数股东和秘鲁税费使合并EBITDA不能直接按集团股本估值。",
        ],
    }


def build_payload() -> dict[str, Any]:
    workbook_contract = _reference_workbook_contract()
    companies = [_zijin_model(), _cmoc_model(), _mmg_model()]
    valuation_policy = {}
    for company in companies:
        valuation = company["valuation"]
        pe = valuation["normalized_pe"]
        pb = valuation["pb_roe"]
        dcf = valuation["fcfe_dcf"]
        valuation_policy[company["company"]] = {
            "normalized_year": valuation["normalized_year"],
            "normalized_pe": {
                "multiple_low": pe["multiple_low"],
                "multiple_high": pe["multiple_high"],
                "parameter_basis": pe["parameter_basis"],
            },
            "pb_roe": {
                "sustainable_roe_low": pb["sustainable_roe_low"],
                "sustainable_roe_high": pb["sustainable_roe_high"],
                "cost_of_equity": pb["cost_of_equity"],
                "terminal_growth": pb["terminal_growth"],
                "parameter_basis": pb["parameter_basis"],
            },
            "fcfe_dcf": {
                "low_value_assumptions": dcf["low_value_assumptions"],
                "high_value_assumptions": dcf["high_value_assumptions"],
                "parameter_basis": dcf["parameter_basis"],
            },
        }
    inputs = {
        "model_version": "copper.independent_model.v2",
        "as_of_date": "2026-07-26",
        "fx_usd_cny": FX_USD_CNY,
        "price_scenarios_usd_t": PRICE_SCENARIOS,
        "reference_workbook": workbook_contract,
        "valuation_policy": valuation_policy,
        "source_policy": (
            "只使用官方财报、官方产量指引和截至2026-07-26的官方经营更新；"
            "本冻结版本未读取卖方盈利预测、Wind一致预期或当前市值。"
        ),
    }
    outputs = {"companies": companies}
    return {
        "schema_version": "copper.independent_model.freeze.v2",
        "research_run_ref": "copper_b_20260726_workbook_revision",
        "freeze_status": "frozen_before_external_reconciliation",
        "inputs": inputs,
        "outputs": outputs,
        "input_sha256": _sha256(inputs),
        "output_sha256": _sha256(outputs),
        "limitations": [
            "已校验并使用《碳酸锂标的估值测算20260606.xlsx》的公式结构，"
            "但不复制锂行业税率、产品折算、成本、税后系数或公司估值倍数。",
            "资源业务隐含市值和单位权益产量市值需要当前市场价格，只能在本独立"
            "模型冻结后的外部对账中计算，不能反向调整冻结预测。",
            "铜价情景是研究输入而非外部事实；每家公司均保留下行、基准、上行情景。",
            "后续一致预期和市场价格只用于对账，不回写或覆盖本冻结文件。",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="构建并冻结铜行业三家公司独立财务模型")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    output = Path(args.output).resolve()
    payload = build_payload()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": str(output),
                "input_sha256": payload["input_sha256"],
                "output_sha256": payload["output_sha256"],
                "companies": [item["company"] for item in payload["outputs"]["companies"]],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
