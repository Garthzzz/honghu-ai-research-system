from __future__ import annotations

"""Build and freeze the lithium/carbonate and company independent models.

This module is deliberately offline.  It reads only:

* the user-provided valuation workbook (formula contract);
* the already captured 2025 Wind actual financial statements; and
* project/operating assumptions transcribed from primary filings and industry
  research before any consensus values are opened.

It does not read the current market snapshot or the Wind/Tushare consensus
sections.  External reconciliation is a separate, post-freeze step.
"""

import argparse
import hashlib
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CACHE = ROOT / "cache" / "lithium_research"
MODEL_DIR = CACHE / "models"
FREEZE_DIR = MODEL_DIR / "company_freezes"
SNAPSHOT_PATH = CACHE / "lithium_financial_snapshot.json"
WORKBOOK_PATH = ROOT / "碳酸锂标的估值测算20260606.xlsx"
DEFAULT_OUTPUT = MODEL_DIR / "lithium_company_independent_models_v1.json"

MODEL_AS_OF = "2026-07-27"
FORECAST_YEARS = (2026, 2027, 2028)
VAT_RATE = 0.13

PRICE_SCENARIOS = {
    "下行情景": {2026: 10.0, 2027: 9.0, 2028: 9.0},
    "基准情景": {2026: 15.0, 2027: 14.0, 2028: 13.0},
    "上行情景": {2026: 18.0, 2027: 17.0, 2028: 16.0},
}


def _canonical_json(payload: Any) -> str:
    return json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _sha256(payload: Any) -> str:
    return "sha256:" + hashlib.sha256(_canonical_json(payload).encode("utf-8")).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return "sha256:" + digest.hexdigest()


def _finite(value: Any, name: str) -> float:
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{name} 必须是有限数")
    return number


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _workbook_contract() -> dict[str, Any]:
    if not WORKBOOK_PATH.is_file():
        raise FileNotFoundError(WORKBOOK_PATH)
    workbook = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    required = {
        "持仓",
        "Sheet1",
        "汇总",
        "赣锋锂业",
        "天齐锂业",
        "中矿资源",
        "盛新锂能",
        "永兴材料",
        "大中矿业",
        "国城矿业",
        "盐湖股份",
        "藏格矿业",
        "华友钴业",
        "雅化集团",
    }
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"估值工作簿缺少工作表: {missing}")
    formulas = {
        "project_attributable_output": workbook["赣锋锂业"]["E5"].value,
        "resource_profit_grid": workbook["赣锋锂业"]["E56"].value,
        "other_business_valuation": workbook["赣锋锂业"]["E78"].value,
        "implied_resource_equity_value": workbook["赣锋锂业"]["E83"].value,
        "implied_resource_pe": workbook["赣锋锂业"]["E84"].value,
        "cross_company_resource_value": workbook["汇总"]["D47"].value,
        "cross_company_resource_multiple": workbook["汇总"]["D58"].value,
    }
    if not all(isinstance(value, str) and value.startswith("=") for value in formulas.values()):
        raise ValueError("估值工作簿的关键公式合同不完整")
    return {
        "path": WORKBOOK_PATH.relative_to(ROOT).as_posix(),
        "sha256": _sha256_file(WORKBOOK_PATH),
        "formula_contract": formulas,
        "transfer_policy": (
            "迁移逐项目权益产量、商品价格利润矩阵、其他业务拆分、资源业务隐含估值和"
            "单位权益产量估值；不照搬工作簿中的旧股价、旧股本、外部链接、锂价、成本、"
            "税后系数或估值倍数。所有这些输入在本模型中重新列示并冻结。"
        ),
    }


def _snapshot_actuals() -> dict[str, dict[str, float]]:
    payload = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))
    annual_2025 = payload["wind"]["annual"]["2025"]
    result: dict[str, dict[str, float]] = {}
    for batch in annual_2025:
        for ticker, row in batch["rows"].items():
            result[ticker] = {
                "revenue_rmb_bn": _finite(row["oper_rev"], f"{ticker}.revenue") / 1e9,
                "net_income_rmb_bn": _finite(
                    row["np_belongto_parcomsh"], f"{ticker}.net_income"
                )
                / 1e9,
                "operating_cash_flow_rmb_bn": _finite(
                    row["net_cash_flows_oper_act"], f"{ticker}.ocf"
                )
                / 1e9,
                "capex_rmb_bn": _finite(
                    row["cash_pay_acq_const_fiolta"], f"{ticker}.capex"
                )
                / 1e9,
                "equity_rmb_bn": _finite(row["tot_equity"], f"{ticker}.equity") / 1e9,
                "assets_rmb_bn": _finite(row["tot_assets"], f"{ticker}.assets") / 1e9,
                "roe_pct": _finite(row["roe"], f"{ticker}.roe"),
                "roa_pct": _finite(row["roa2"], f"{ticker}.roa"),
            }
    if len(result) != 13:
        raise ValueError(f"2025实际财务应覆盖13家公司，实际为{len(result)}")
    return result


def _build_lithium_supply_demand() -> dict[str, Any]:
    """Global primary-lithium balance in million tonnes LCE.

    The former implementation rebuilt regional and end-use components from
    mutually incompatible broker definitions.  That made the totals look
    precise while moving the global balance into deficit several years earlier
    than the latest official reference.  The live baseline now follows one
    internally consistent table from the Australian Government's June 2026
    Resources and Energy Quarterly.  Country mine production and company
    chemical-sales concentration are retained as separate diagnostic ledgers.
    """
    official_balance = {
        2025: {"supply": 1.491, "demand": 1.376},
        2026: {"supply": 1.716, "demand": 1.572},
        2027: {"supply": 1.945, "demand": 1.795},
        2028: {"supply": 2.150, "demand": 2.033},
        2029: {"supply": 2.276, "demand": 2.229},
        2030: {"supply": 2.379, "demand": 2.411},
        2031: {"supply": 2.636, "demand": 2.641},
    }
    base_rows: list[dict[str, Any]] = []
    for year, values in official_balance.items():
        supply_total = values["supply"]
        demand_total = values["demand"]
        base_rows.append(
            {
                "year": year,
                "available_supply_mt_lce": supply_total,
                "demand_mt_lce": demand_total,
                "balance_mt_lce": round(supply_total - demand_total, 3),
                # Compatibility fields are intentionally single-line totals.
                # Regional and end-use estimates with a different statistical
                # scope must not be forced to add to this official series.
                "supply_components": {"澳大利亚政府全球供给基准": supply_total},
                "demand_components": {"澳大利亚政府全球需求基准": demand_total},
                "status": (
                    "澳大利亚政府2026年6月公布的2025年估计值"
                    if year == 2025
                    else "澳大利亚政府2026年6月预测"
                ),
            }
        )
    stress_ratios = {
        "投产顺利情景": {
            "supply": {
                2026: 0.03, 2027: 0.05, 2028: 0.07,
                2029: 0.08, 2030: 0.09, 2031: 0.10,
            },
            "demand": {
                2026: -0.02, 2027: -0.03, 2028: -0.04,
                2029: -0.04, 2030: -0.05, 2031: -0.05,
            },
        },
        "基准情景": {
            "supply": {year: 0.0 for year in range(2026, 2032)},
            "demand": {year: 0.0 for year in range(2026, 2032)},
        },
        "投产受限且需求偏强情景": {
            "supply": {
                2026: -0.03, 2027: -0.05, 2028: -0.07,
                2029: -0.09, 2030: -0.10, 2031: -0.10,
            },
            "demand": {
                2026: 0.02, 2027: 0.03, 2028: 0.04,
                2029: 0.05, 2030: 0.05, 2031: 0.05,
            },
        },
    }
    by_year = {row["year"]: row for row in base_rows}
    scenarios: dict[str, list[dict[str, Any]]] = {}
    for name, delta in stress_ratios.items():
        rows = []
        for year in range(2026, 2032):
            supply = by_year[year]["available_supply_mt_lce"] * (
                1.0 + delta["supply"][year]
            )
            demand_value = by_year[year]["demand_mt_lce"] * (
                1.0 + delta["demand"][year]
            )
            rows.append(
                {
                    "year": year,
                    "available_supply_mt_lce": round(supply, 3),
                    "demand_mt_lce": round(demand_value, 3),
                    "balance_mt_lce": round(supply - demand_value, 3),
                    "balance_ratio_pct": round((supply - demand_value) / demand_value * 100, 2),
                }
            )
        scenarios[name] = rows
    payload = {
        "schema_version": "lithium_supply_demand_model.v1",
        "as_of_date": MODEL_AS_OF,
        "unit": "百万吨LCE",
        "scope": (
            "全球锂生产与消费折合LCE，严格沿用澳大利亚政府Resources and "
            "Energy Quarterly 2026年6月表格口径；不是中国碳酸锂产品平衡，"
            "也不能与卖方含回收/在途/可销售产品的宽口径直接拼接。"
        ),
        "method": (
            "基准序列直接采用同一官方表中的2025年估计和2026—2031年预测。"
            "压力测试在官方总量上分别调整项目兑现率和需求强弱，不改变基准事实；"
            "供给、需求和余额均保留三位小数，避免不同来源的分项被强行勾稽。"
        ),
        "base_rows": base_rows,
        "scenarios": scenarios,
        "stress_ratio_inputs": stress_ratios,
        "country_mine_2025": {
            "unit": "千吨锂金属",
            "rows": {
                "澳大利亚": 92.0,
                "中国": 62.0,
                "智利": 56.0,
                "津巴布韦": 28.0,
                "阿根廷": 23.0,
                "巴西": 12.0,
                "马里": 9.4,
                "加拿大": 5.6,
                "葡萄牙": 0.38,
            },
            "covered_total_kt_li": 288.38,
            "cr3_pct": 72.41,
            "cr5_pct": 90.00,
            "source": "USGS Mineral Commodity Summaries 2026",
            "scope_note": "美国矿山产量未披露；CR3/CR5按USGS已列国家和世界29万吨分母计算。",
        },
        "external_reference_ranges": [
            {
                "publisher": "东吴证券",
                "date": "2026-02-26",
                "supply_mt_lce": {2026: 2.14, 2027: 2.61, 2028: 3.00},
                "demand_mt_lce": {2026: 2.10, 2027: 2.50, 2028: 2.85},
                "use": "卖方宽口径情景对照；先核对回收、产品和在途口径，不与官方序列取平均",
            },
            {
                "publisher": "浙商证券",
                "date": "2026-06-14",
                "supply_mt_lce": {2026: 2.25, 2027: 2.61, 2028: 3.02},
                "demand_mt_lce": {2026: 2.16, 2027: 2.70, 2028: 3.34},
                "use": "卖方宽口径情景对照；先核对回收、产品和在途口径，不与官方序列取平均",
            },
        ],
        "key_limits": [
            "官方表的全球总量与USGS矿端、SQM化学品销售份额是三个不同分母。",
            "压力测试是研究假设，不是第二套官方预测，也不是概率加权结果。",
            "2028年以后结果对非洲新矿、阿根廷盐湖、中国回收和储能需求最敏感。",
            "供需差不足以单独决定价格；库存、长协、产品结构和政策会改变价格传导速度。",
        ],
    }
    payload["content_sha256"] = _sha256(payload)
    return payload


def _build_carbonate_supply_demand() -> dict[str, Any]:
    """China lithium-carbonate product balance, separate from global LCE."""
    rows = [
        {
            "year": 2025,
            "domestic_output_mt": 0.976,
            "imports_mt": 0.2430,
            "exports_mt": 0.005290,
            "available_supply_mt": 1.213710,
            "demand_mt": 1.212,
            "balance_mt": 0.001710,
            "status": "产量和海关为实际值；需求为外部行业估计，余额不等于库存变动",
        },
        {
            "year": 2026,
            "domestic_output_mt": 1.395,
            "imports_mt": 0.365,
            "exports_mt": 0.005,
            "available_supply_mt": 1.755,
            "demand_mt": 1.770,
            "balance_mt": -0.015,
            "status": "全年研究中值；上半年实际区间和下半年公开展望为约束",
        },
        {
            "year": 2027,
            "domestic_output_mt": 1.550,
            "imports_mt": 0.390,
            "exports_mt": 0.008,
            "available_supply_mt": 1.932,
            "demand_mt": 1.950,
            "balance_mt": -0.018,
            "status": "研究中值，结果必须结合宽区间而非单点解读",
        },
        {
            "year": 2028,
            "domestic_output_mt": 1.720,
            "imports_mt": 0.430,
            "exports_mt": 0.010,
            "available_supply_mt": 2.140,
            "demand_mt": 2.140,
            "balance_mt": 0.000,
            "status": "研究中值，方向取决于国内项目、进口和储能需求同时兑现",
        },
    ]
    for row in rows:
        expected_supply = row["domestic_output_mt"] + row["imports_mt"] - row["exports_mt"]
        if abs(expected_supply - row["available_supply_mt"]) > 1e-9:
            raise ValueError(f"{row['year']}年碳酸锂可用供给无法勾稽")
        if abs(row["available_supply_mt"] - row["demand_mt"] - row["balance_mt"]) > 1e-9:
            raise ValueError(f"{row['year']}年碳酸锂平衡无法勾稽")
        row["balance_ratio_pct"] = round(row["balance_mt"] / row["demand_mt"] * 100, 2)
    payload = {
        "schema_version": "lithium_carbonate_supply_demand_model.v1",
        "as_of_date": MODEL_AS_OF,
        "unit": "百万吨碳酸锂",
        "scope": "中国碳酸锂产品表观平衡；与全球锂资源LCE平衡分开，不能相互替代。",
        "formula": "可用供给＝国内产量＋进口－出口；产品平衡＝可用供给－下游需求。",
        "rows": rows,
        "observed_2026_h1": {
            "period": "2026H1",
            "domestic_output_range_mt": [0.5936, 0.6300],
            "smm_domestic_output_mt": 0.6220,
            "imports_mt": 0.1790,
            "exports_mt": 0.002348,
            "apparent_supply_range_mt": [0.770252, 0.806652],
            "route_output_mt": {
                "锂辉石": 0.3737,
                "锂云母": 0.0767,
                "盐湖": 0.0905,
                "回收": 0.0545,
            },
            "demand_monthly_start_end_mt": [0.1247, 0.1510],
            "inventory_sample_mt": 0.0977,
            "note": (
                "隆众、Mysteel与SMM统计覆盖不同，因此保留区间；"
                "表观供给未扣库存变化，月度需求只保留已公开的起止点。"
            ),
        },
        "forecast_ranges": {
            "2026": {
                "domestic_output_mt": [1.38, 1.41],
                "net_imports_mt": [0.35, 0.37],
                "demand_mt": [1.72, 1.80],
                "balance_mt": [-0.07, 0.06],
            },
            "2027": {
                "domestic_output_mt": [1.45, 1.65],
                "net_imports_mt": [0.36, 0.42],
                "demand_mt": [1.85, 2.05],
                "balance_mt": [-0.24, 0.22],
            },
            "2028": {
                "domestic_output_mt": [1.58, 1.82],
                "net_imports_mt": [0.38, 0.46],
                "demand_mt": [2.00, 2.25],
                "balance_mt": [-0.29, 0.28],
            },
        },
        "evidence_constraints": [
            "锂业分会披露2025年国内碳酸锂产量97.6万吨、净进口23.7万吨。",
            "海关口径2025年进口24.30万吨、出口0.529万吨。",
            "2026年上半年进口17.9万吨、出口0.2348万吨；产量口径为59.36万—63.0万吨。",
            "SMM口径上半年产量约62.2万吨，月度需求从12.47万吨升至15.10万吨。",
            "库存样本扩容会制造表观跳变，因此库存只用于方向核验，不直接拼入供给。",
        ],
        "key_limits": [
            "2025需求来自行业数据库估计，与产量和海关并非同一统计机构，余额只做方向检查。",
            "2026—2028年度数值是研究中值；应优先阅读forecast_ranges而不是把中值当确定预测。",
            "进口存在保税、转口和锂盐品类口径差异。",
            "部分氢氧化锂可转化为碳酸锂，但转换成本和时滞使二者不是完全同质供给。",
            "年度小缺口可由库存满足，不代表现货价格必然单边上涨。",
        ],
    }
    payload["content_sha256"] = _sha256(payload)
    return payload


COMPANIES: dict[str, dict[str, Any]] = {
    "赣锋锂业": {
        "company_id": 640,
        "ticker": "002460.SZ",
        "model_type": "全球资源与锂盐一体化",
        "product_volume": {2026: 12.5, 2027: 16.0, 2028: 19.0},
        "resource_share": {2026: 0.80, 2027: 0.82, 2028: 0.84},
        "resource_cost": {2026: 6.0, 2027: 5.8, 2028: 5.7},
        "processing_margin": {2026: 1.3, 2027: 1.3, 2028: 1.4},
        "after_tax_factor": 0.70,
        "other_profit": {2026: 0.50, 2027: 0.70, 2028: 0.90},
        "corporate_cost": {2026: 0.20, 2027: 0.25, 2028: 0.30},
        "revenue": {2026: 38.0, 2027: 46.0, 2028: 54.0},
        "equity": {2026: 55.0, 2027: 60.0, 2028: 67.0},
        "fcfe_conversion": 0.68,
        "pe_range": (12.0, 16.0),
        "pb_haircut": 0.85,
        "valuation_note": "资源、盐湖、锂盐及电池业务并存，PE、PB—ROE和FCFE交叉验证。",
        "source_facts": [
            "Cauchari-Olaroz由公司控制46.67%，2025年产量3.41万吨碳酸锂，2026年目标3.5—4.0万吨。",
            "Goulamina由公司控制65%，2025年干基精矿产量33.66万吨；Mariana 2万吨氯化锂产线已投产。",
            "加不斯一期60万吨选矿项目处于爬坡，四川赣锋5万吨锂盐项目进入爬坡。",
        ],
    },
    "融捷股份": {
        "company_id": 641,
        "ticker": "002192.SZ",
        "model_type": "高品位锂矿与权益锂盐",
        "product_volume": {2026: 2.32, 2027: 2.60, 2028: 3.60},
        "resource_share": {2026: 1.00, 2027: 1.00, 2028: 1.00},
        "resource_cost": {2026: 4.8, 2027: 4.7, 2028: 4.7},
        "processing_margin": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "after_tax_factor": 0.70,
        "other_profit": {2026: 0.08, 2027: 0.10, 2028: 0.12},
        "corporate_cost": {2026: 0.20, 2027: 0.21, 2028: 0.23},
        "revenue": {2026: 1.5, 2027: 1.8, 2028: 2.4},
        "equity": {2026: 4.5, 2027: 5.5, 2028: 7.0},
        "fcfe_conversion": 0.70,
        "pe_range": (10.0, 14.0),
        "pb_haircut": 0.75,
        "valuation_note": "产量集中且扩产审批敏感，资源利润法为主，PB—ROE只作周期约束。",
        "source_facts": [
            "134号脉采矿能力105万吨/年、选矿能力45万吨/年，另规划35万吨/年选矿扩建。",
            "2025年锂精矿产量18.56万吨，同比增长174.83%。",
            "合并锂盐产能0.3万吨，联营2万吨锂盐产线满产并规划扩至4万吨。",
        ],
    },
    "盛新锂能": {
        "company_id": 642,
        "ticker": "002240.SZ",
        "model_type": "矿盐一体化并含外购矿加工",
        "product_volume": {2026: 10.0, 2027: 12.0, 2028: 13.5},
        "resource_share": {2026: 0.40, 2027: 0.46, 2028: 0.52},
        "resource_cost": {2026: 6.2, 2027: 6.0, 2028: 5.9},
        "processing_margin": {2026: 1.2, 2027: 1.3, 2028: 1.4},
        "after_tax_factor": 0.68,
        "other_profit": {2026: 0.05, 2027: 0.08, 2028: 0.10},
        "corporate_cost": {2026: 0.75, 2027: 0.75, 2028: 0.78},
        "revenue": {2026: 12.0, 2027: 14.5, 2028: 16.5},
        "equity": {2026: 13.0, 2027: 15.0, 2028: 17.5},
        "fcfe_conversion": 0.58,
        "pe_range": (10.0, 14.0),
        "pb_haircut": 0.75,
        "valuation_note": "外购矿与自有矿成本差异大，必须拆分资源利润和加工利润。",
        "source_facts": [
            "已建锂盐产能13.7万吨/年、锂金属500吨/年。",
            "业隆沟精矿产能7.5万吨/年，Sabi Star精矿产能29万吨/年。",
            "木绒矿设计原矿处理能力300万吨/年，投产和爬坡决定自给率改善速度。",
        ],
    },
    "盐湖股份": {
        "company_id": 643,
        "ticker": "000792.SZ",
        "model_type": "低成本盐湖锂与钾肥",
        "product_volume": {2026: 6.5, 2027: 8.0, 2028: 8.0},
        "resource_share": {2026: 1.00, 2027: 1.00, 2028: 1.00},
        "resource_cost": {2026: 3.3, 2027: 3.2, 2028: 3.2},
        "processing_margin": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "after_tax_factor": 0.72,
        "other_profit": {2026: 5.80, 2027: 6.00, 2028: 6.20},
        "corporate_cost": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "revenue": {2026: 23.0, 2027: 26.0, 2028: 28.0},
        "equity": {2026: 55.0, 2027: 65.0, 2028: 75.0},
        "fcfe_conversion": 0.88,
        "pe_range": (11.0, 14.0),
        "pb_haircut": 0.90,
        "valuation_note": "钾肥提供利润底仓，锂业务用资源利润法，PB—ROE适用性高于纯加工企业。",
        "source_facts": [
            "原有碳酸锂设计产能4万吨/年，2025年维持高利用率。",
            "新建4万吨/年一体化项目进入试生产。",
            "2025年碳酸锂产量约4.65万吨、销量约4.56万吨，钾肥仍是利润底仓。",
        ],
    },
    "大中矿业": {
        "company_id": 644,
        "ticker": "001203.SZ",
        "model_type": "铁矿现金流加锂项目期权",
        "product_volume": {2026: 0.5, 2027: 3.0, 2028: 6.0},
        "resource_share": {2026: 1.00, 2027: 1.00, 2028: 1.00},
        "resource_cost": {2026: 4.8, 2027: 4.7, 2028: 4.6},
        "processing_margin": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "after_tax_factor": 0.68,
        "other_profit": {2026: 0.80, 2027: 0.85, 2028: 0.90},
        "corporate_cost": {2026: 0.08, 2027: 0.10, 2028: 0.12},
        "revenue": {2026: 6.0, 2027: 9.0, 2028: 12.0},
        "equity": {2026: 8.5, 2027: 10.0, 2028: 13.5},
        "fcfe_conversion": 0.45,
        "pe_range": (9.0, 13.0),
        "pb_haircut": 0.70,
        "valuation_note": "铁矿是当前利润基础，锂项目用分部估值并对投产延迟做折价。",
        "source_facts": [
            "鸡脚山一期2万吨碳酸锂项目计划2026年投产。",
            "加达锂矿预计2027年开始贡献。",
            "铁矿业务是当前现金流基础，锂业务在投产前不应替代现有利润。",
        ],
    },
    "雅化集团": {
        "company_id": 645,
        "ticker": "002497.SZ",
        "model_type": "自有矿加锂盐与民爆",
        "product_volume": {2026: 10.0, 2027: 11.0, 2028: 12.0},
        "resource_share": {2026: 0.35, 2027: 0.40, 2028: 0.44},
        "resource_cost": {2026: 6.5, 2027: 6.2, 2028: 6.0},
        "processing_margin": {2026: 0.9, 2027: 1.0, 2028: 1.1},
        "after_tax_factor": 0.68,
        "other_profit": {2026: 0.35, 2027: 0.38, 2028: 0.40},
        "corporate_cost": {2026: 0.15, 2027: 0.17, 2028: 0.19},
        "revenue": {2026: 15.0, 2027: 18.0, 2028: 21.0},
        "equity": {2026: 12.5, 2027: 14.5, 2028: 17.0},
        "fcfe_conversion": 0.62,
        "pe_range": (10.0, 14.0),
        "pb_haircut": 0.78,
        "valuation_note": "矿盐一体化改善成本，民爆提供非锂利润；需拆分外购矿加工。",
        "source_facts": [
            "Kamativi精矿产能35万吨/年，已进入规模化生产。",
            "李家沟精矿产能18—20万吨/年，2025年9月正式投产并达产。",
            "锂盐设计产能约13万吨/年，民爆业务提供非锂利润。",
        ],
    },
    "天华新能": {
        "company_id": 646,
        "ticker": "300390.SZ",
        "model_type": "大规模锂盐加工与非洲资源",
        "product_volume": {2026: 13.0, 2027: 18.0, 2028: 22.0},
        "resource_share": {2026: 0.55, 2027: 0.62, 2028: 0.68},
        "resource_cost": {2026: 7.0, 2027: 6.7, 2028: 6.5},
        "processing_margin": {2026: 1.1, 2027: 1.2, 2028: 1.3},
        "after_tax_factor": 0.68,
        "other_profit": {2026: 0.20, 2027: 0.22, 2028: 0.24},
        "corporate_cost": {2026: 0.25, 2027: 0.30, 2028: 0.35},
        "revenue": {2026: 18.0, 2027: 23.0, 2028: 27.0},
        "equity": {2026: 18.0, 2027: 22.0, 2028: 27.0},
        "fcfe_conversion": 0.50,
        "pe_range": (10.0, 15.0),
        "pb_haircut": 0.78,
        "valuation_note": "资源自给率决定利润弹性；扩产期资本开支使FCFE权重低于PE。",
        "source_facts": [
            "已建锂盐产能16.5万吨/年，其中氢氧化锂13.5万吨、碳酸锂3万吨。",
            "2025年氢氧化锂产量5.95万吨、碳酸锂产量4.74万吨。",
            "Ogapa矿公司经济权益37.5%，金子峰项目计划2027年贡献。",
        ],
    },
    "天齐锂业": {
        "company_id": 647,
        "ticker": "002466.SZ",
        "model_type": "Greenbushes资源、锂盐与SQM权益",
        "product_volume": {2026: 7.5, 2027: 8.5, 2028: 9.2},
        "resource_share": {2026: 0.88, 2027: 0.90, 2028: 0.90},
        "resource_cost": {2026: 4.5, 2027: 4.4, 2028: 4.4},
        "processing_margin": {2026: 0.8, 2027: 0.9, 2028: 0.9},
        "after_tax_factor": 0.67,
        "other_profit": {2026: 1.00, 2027: 1.30, 2028: 1.50},
        "corporate_cost": {2026: 0.60, 2027: 0.60, 2028: 0.60},
        "revenue": {2026: 22.0, 2027: 25.0, 2028: 27.0},
        "equity": {2026: 58.0, 2027: 64.0, 2028: 71.0},
        "fcfe_conversion": 0.72,
        "pe_range": (11.0, 15.0),
        "pb_haircut": 0.82,
        "valuation_note": "Greenbushes、自产锂盐与SQM权益法收益分开核验，不能把SQM销量重复并入自产。",
        "source_facts": [
            "Greenbushes五座选矿厂合计精矿产能214万吨/年。",
            "2025年Greenbushes精矿产量135万吨，其中化学级130万吨；CGP3于2026年1月产出首批产品。",
            "公司对Greenbushes的穿透权益与SQM权益法收益均需单独处理，不能按100%项目产量估值。",
        ],
    },
    "永杉锂业": {
        "company_id": 648,
        "ticker": "603399.SH",
        "model_type": "外购原料锂盐加工",
        "product_volume": {2026: 4.2, 2027: 6.0, 2028: 6.5},
        "resource_share": {2026: 0.00, 2027: 0.00, 2028: 0.00},
        "resource_cost": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "processing_margin": {2026: 1.0, 2027: 1.1, 2028: 1.2},
        "after_tax_factor": 0.65,
        "other_profit": {2026: 0.02, 2027: 0.03, 2028: 0.04},
        "corporate_cost": {2026: 0.25, 2027: 0.27, 2028: 0.29},
        "revenue": {2026: 8.0, 2027: 10.0, 2028: 11.0},
        "equity": {2026: 1.5, 2027: 2.0, 2028: 2.8},
        "fcfe_conversion": 0.55,
        "pe_range": (8.0, 12.0),
        "pb_haircut": 0.55,
        "valuation_note": "以加工费和库存周期建模，不把外购原料销量当资源租金。",
        "source_facts": [
            "2025年碳酸锂产量1.3814万吨、氢氧化锂产量1.4463万吨。",
            "锂盐产能由2.5万吨增至4.5万吨。",
            "2.2万吨扩建目标2026年10月投产，届时总产能超过6万吨。",
        ],
    },
    "中矿资源": {
        "company_id": 649,
        "ticker": "002738.SZ",
        "model_type": "非洲锂矿、锂盐与铯铷",
        "product_volume": {2026: 7.0, 2027: 7.5, 2028: 8.0},
        "resource_share": {2026: 0.85, 2027: 0.88, 2028: 0.90},
        "resource_cost": {2026: 5.5, 2027: 5.3, 2028: 5.2},
        "processing_margin": {2026: 1.0, 2027: 1.1, 2028: 1.2},
        "after_tax_factor": 0.68,
        "other_profit": {2026: 0.30, 2027: 0.35, 2028: 0.40},
        "corporate_cost": {2026: 0.30, 2027: 0.32, 2028: 0.34},
        "revenue": {2026: 10.0, 2027: 13.5, 2028: 16.0},
        "equity": {2026: 14.5, 2027: 17.5, 2028: 21.0},
        "fcfe_conversion": 0.62,
        "pe_range": (11.0, 15.0),
        "pb_haircut": 0.82,
        "valuation_note": "资源自给率高，铯铷提供非锂利润；PE、SOTP和PB—ROE均可使用。",
        "source_facts": [
            "Bikita锂资源量折合343.41万吨LCE。",
            "两条200万吨/年选矿线分别生产锂辉石和透锂长石精矿，设计精矿各约30万吨。",
            "锂盐产能7.1万吨/年，升级后的3万吨产线于2026年1月投产。",
        ],
    },
    "藏格矿业": {
        "company_id": 650,
        "ticker": "000408.SZ",
        "model_type": "盐湖锂、钾肥与巨龙铜业权益",
        "product_volume": {2026: 1.72, 2027: 3.75, 2028: 4.10},
        "resource_share": {2026: 1.00, 2027: 1.00, 2028: 1.00},
        "resource_cost": {2026: 3.8, 2027: 4.0, 2028: 4.0},
        "processing_margin": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "after_tax_factor": 0.70,
        "other_profit": {2026: 5.70, 2027: 6.60, 2028: 7.20},
        "corporate_cost": {2026: 0.10, 2027: 0.12, 2028: 0.15},
        "revenue": {2026: 5.0, 2027: 6.5, 2028: 8.0},
        "equity": {2026: 20.0, 2027: 25.0, 2028: 31.0},
        "fcfe_conversion": 0.82,
        "pe_range": (12.0, 16.0),
        "pb_haircut": 0.88,
        "valuation_note": "钾肥和巨龙铜业决定利润底仓，锂项目只对增量估值。",
        "source_facts": [
            "2025年碳酸锂产量0.8808万吨、销量0.8957万吨。",
            "原有设计产能1万吨/年，2025年利用率88.08%。",
            "Mamico 5万吨项目公司间接权益26.95%，目标2026年下半年开始贡献；钾肥和巨龙铜业是重要利润来源。",
        ],
    },
    "西藏城投": {
        "company_id": 651,
        "ticker": "600773.SH",
        "model_type": "未商业化盐湖项目与地产存量",
        "product_volume": {2026: 0.00, 2027: 0.08, 2028: 0.16},
        "resource_share": {2026: 1.00, 2027: 1.00, 2028: 1.00},
        "resource_cost": {2026: 5.5, 2027: 5.5, 2028: 5.3},
        "processing_margin": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "after_tax_factor": 0.65,
        "other_profit": {2026: -0.20, 2027: -0.10, 2028: 0.0},
        "corporate_cost": {2026: 0.10, 2027: 0.10, 2028: 0.10},
        "revenue": {2026: 0.8, 2027: 1.0, 2028: 1.2},
        "equity": {2026: 3.6, 2027: 3.5, 2028: 3.5},
        "fcfe_conversion": 0.0,
        "net_income_override": {2026: -0.30, 2027: -0.15, 2028: -0.05},
        "net_income_override_rationale": (
            "盐湖项目尚未商业化，规划产能与小试/中试产量不计入基准归母利润；"
            "2026—2028使用地产存量和公司层损益逐步收敛的独立亏损路径。"
        ),
        "pe_range": None,
        "pb_haircut": 0.45,
        "valuation_note": "项目尚未商业化，不使用PE；只做条件化项目NAV和资产负债约束。",
        "source_facts": [
            "公司持有国能矿业41%权益。",
            "龙木错与结则茶卡合计资源量约390万吨LCE。",
            "项目尚未商业化，7万吨和3万吨规划不能当作当前产量或基准利润。",
        ],
    },
    "永兴材料": {
        "company_id": 652,
        "ticker": "002756.SZ",
        "model_type": "云母矿盐一体化与特钢",
        "product_volume": {2026: 2.9, 2027: 4.2, 2028: 5.8},
        "resource_share": {2026: 1.00, 2027: 1.00, 2028: 1.00},
        "resource_cost": {2026: 5.5, 2027: 5.3, 2028: 5.2},
        "processing_margin": {2026: 0.0, 2027: 0.0, 2028: 0.0},
        "after_tax_factor": 0.70,
        "other_profit": {2026: 0.30, 2027: 0.32, 2028: 0.35},
        "corporate_cost": {2026: 0.10, 2027: 0.11, 2028: 0.12},
        "revenue": {2026: 9.0, 2027: 10.5, 2028: 12.0},
        "equity": {2026: 14.0, 2027: 16.0, 2028: 19.0},
        "fcfe_conversion": 0.78,
        "pe_range": (9.0, 13.0),
        "pb_haircut": 0.78,
        "valuation_note": "锂云母成本曲线和环保约束重要，特钢提供非锂利润。",
        "source_facts": [
            "2025年锂业务产量2.4823万吨。",
            "公司拥有较大锂云母资源并形成采选冶一体化。",
            "特钢业务提供非锂利润和现金流缓冲。",
        ],
    },
}


def _resource_profit(
    *,
    volume_10kt: float,
    resource_share: float,
    price_10k_rmb_t: float,
    resource_cost_10k_rmb_t: float,
    processing_margin_10k_rmb_t: float,
    after_tax_factor: float,
) -> dict[str, float]:
    resource_volume = volume_10kt * resource_share
    processing_volume = volume_10kt - resource_volume
    resource_after_tax = (
        resource_volume
        * (price_10k_rmb_t - resource_cost_10k_rmb_t)
        / (1.0 + VAT_RATE)
        * after_tax_factor
        / 10.0
    )
    processing_after_tax = (
        processing_volume
        * processing_margin_10k_rmb_t
        / (1.0 + VAT_RATE)
        * after_tax_factor
        / 10.0
    )
    return {
        "resource_volume_10kt_lce": resource_volume,
        "processing_volume_10kt_lce": processing_volume,
        "resource_after_tax_profit_rmb_bn": resource_after_tax,
        "processing_after_tax_profit_rmb_bn": processing_after_tax,
    }


def _fcfe_value(fcfe: list[float], cost_of_equity: float, terminal_growth: float) -> dict[str, float]:
    if len(fcfe) != 3 or not 0 <= terminal_growth < cost_of_equity:
        raise ValueError("FCFE估值参数不合法")
    discounted = [
        value / ((1.0 + cost_of_equity) ** index)
        for index, value in enumerate(fcfe, start=1)
    ]
    terminal = fcfe[-1] * (1.0 + terminal_growth) / (cost_of_equity - terminal_growth)
    terminal_pv = terminal / ((1.0 + cost_of_equity) ** 3)
    equity_value = sum(discounted) + terminal_pv
    return {
        "equity_value_rmb_bn": equity_value,
        "terminal_value_share_pct": terminal_pv / equity_value * 100,
    }


def _pe_parameter_basis(name: str, spec: dict[str, Any]) -> str:
    common = (
        "行业倍数锚来自东吴证券2026-02-26碳酸锂行业专题：2026年权益资源"
        "利润使用10—15倍PE、远期满产利润使用更低倍数；本模型选择2027年"
        "正常化集团利润，避免把项目价值重复加入。"
    )
    company_adjustments = {
        "赣锋锂业": "全球资源与电池业务提供分散度，上沿保留规模溢价；多项目爬坡与资本开支限制上沿。",
        "融捷股份": "高品位资源支持中部倍数，但单一核心矿山和小市值波动使上限低于一体化龙头。",
        "盛新锂能": "资源覆盖改善但仍含外购矿加工和在建项目，采用行业中低部。",
        "盐湖股份": "低成本盐湖和钾肥底仓提高稳定性，但商品周期与扩建资本使上限不超过14倍。",
        "大中矿业": "铁矿现金流提供底仓，锂项目尚在建设爬坡，使用低于成熟锂资源公司的区间。",
        "雅化集团": "民爆业务提供下行缓冲，锂资源自给仍在提升，采用行业中部。",
        "天华新能": "锂盐规模和客户结构支持上沿，非洲资源兑现与扩张资本限制进一步溢价。",
        "天齐锂业": "Greenbushes和SQM权益支持较高质量倍数，权益穿透、资本结构和周期限制上沿。",
        "永杉锂业": "外购原料加工、库存和价差波动使PE只作低倍数参考，不给予资源溢价。",
        "中矿资源": "资源自给与铯铷利润提高质量，非洲项目、扩产资本和周期限制上沿。",
        "藏格矿业": "钾肥与巨龙铜业权益形成利润底仓，但高权益法利润和高周期ROE不能永久化。",
        "永兴材料": "在产云母资源与特钢形成底仓，云母成本曲线和环保资本约束上沿。",
    }
    return f"{common}{company_adjustments.get(name, spec['valuation_note'])}"


def _company_model(name: str, spec: dict[str, Any], actual: dict[str, float]) -> dict[str, Any]:
    scenarios: dict[str, list[dict[str, Any]]] = {}
    for scenario, prices in PRICE_SCENARIOS.items():
        rows: list[dict[str, Any]] = []
        for year in FORECAST_YEARS:
            bridge = _resource_profit(
                volume_10kt=spec["product_volume"][year],
                resource_share=spec["resource_share"][year],
                price_10k_rmb_t=prices[year],
                resource_cost_10k_rmb_t=spec["resource_cost"][year],
                processing_margin_10k_rmb_t=spec["processing_margin"][year],
                after_tax_factor=spec["after_tax_factor"],
            )
            other_multiplier = (
                0.88 if scenario == "下行情景" else 1.10 if scenario == "上行情景" else 1.0
            )
            other_profit = spec["other_profit"][year] * other_multiplier
            net_income = (
                bridge["resource_after_tax_profit_rmb_bn"]
                + bridge["processing_after_tax_profit_rmb_bn"]
                + other_profit
                - spec["corporate_cost"][year]
            )
            if spec.get("net_income_override"):
                net_income = spec["net_income_override"][year]
            revenue_multiplier = (
                0.78 if scenario == "下行情景" else 1.18 if scenario == "上行情景" else 1.0
            )
            revenue = spec["revenue"][year] * revenue_multiplier
            average_equity = (
                (actual["equity_rmb_bn"] + spec["equity"][year]) / 2.0
                if year == 2026
                else (spec["equity"][year - 1] + spec["equity"][year]) / 2.0
            )
            roe = net_income / average_equity * 100 if average_equity > 0 else None
            rows.append(
                {
                    "year": year,
                    "carbonate_price_10k_rmb_t_incl_vat": prices[year],
                    "product_volume_10kt_lce": spec["product_volume"][year],
                    "resource_share_pct": spec["resource_share"][year] * 100,
                    "resource_cost_10k_rmb_t_incl_vat": spec["resource_cost"][year],
                    "processing_margin_10k_rmb_t_incl_vat": spec["processing_margin"][year],
                    **{key: round(value, 4) for key, value in bridge.items()},
                    "other_profit_rmb_bn": round(other_profit, 4),
                    "corporate_cost_rmb_bn": spec["corporate_cost"][year],
                    "revenue_rmb_bn": round(revenue, 4),
                    "net_income_rmb_bn": round(net_income, 4),
                    "fcfe_rmb_bn": round(net_income * spec["fcfe_conversion"], 4),
                    "equity_rmb_bn": spec["equity"][year],
                    "roe_pct": round(roe, 2) if roe is not None else None,
                }
            )
        scenarios[scenario] = rows

    base = scenarios["基准情景"]
    base_2027 = base[1]
    valuations: list[dict[str, Any]] = []
    if spec["pe_range"] and base_2027["net_income_rmb_bn"] > 0:
        low_pe, high_pe = spec["pe_range"]
        valuations.append(
            {
                "method": "正常化市盈率",
                "role": "核心",
                "forecast_year": 2027,
                "low_rmb_bn": round(base_2027["net_income_rmb_bn"] * low_pe, 2),
                "high_rmb_bn": round(base_2027["net_income_rmb_bn"] * high_pe, 2),
                "inputs": {
                    "net_income_rmb_bn": base_2027["net_income_rmb_bn"],
                    "pe_range": [low_pe, high_pe],
                },
                "parameter_basis": _pe_parameter_basis(name, spec),
            }
        )
    forecast_roe = base[2]["roe_pct"] / 100.0
    if forecast_roe > 0.03:
        sustainable_roe_high = min(0.22, forecast_roe)
        sustainable_roe_low = max(
            0.02, sustainable_roe_high * spec["pb_haircut"]
        )
        pb_low_ke = 0.125
        pb_low_growth = 0.02
        pb_high_ke = 0.105
        pb_high_growth = 0.03
        pb_low = max(
            0.0,
            (sustainable_roe_low - pb_low_growth)
            / (pb_low_ke - pb_low_growth),
        )
        pb_high = max(
            pb_low,
            (sustainable_roe_high - pb_high_growth)
            / (pb_high_ke - pb_high_growth),
        )
        adopted_pb_low = round(pb_low, 2)
        adopted_pb_high = round(pb_high, 2)
        valuations.append(
            {
                "method": "PB—ROE",
                "role": (
                    "有效参考"
                    if name in {"赣锋锂业", "盐湖股份", "天齐锂业", "中矿资源", "永兴材料"}
                    else "诊断"
                ),
                "forecast_year": 2027,
                "low_rmb_bn": round(
                    spec["equity"][2027] * adopted_pb_low, 2
                ),
                "high_rmb_bn": round(
                    spec["equity"][2027] * adopted_pb_high, 2
                ),
                "inputs": {
                    "sustainable_roe_range_pct": [
                        round(sustainable_roe_low * 100, 2),
                        round(sustainable_roe_high * 100, 2),
                    ],
                    "base_cost_of_equity_pct": 11.5,
                    "base_terminal_growth_pct": 2.5,
                    "low_value_cost_of_equity_pct": pb_low_ke * 100,
                    "low_value_terminal_growth_pct": pb_low_growth * 100,
                    "high_value_cost_of_equity_pct": pb_high_ke * 100,
                    "high_value_terminal_growth_pct": pb_high_growth * 100,
                    "roe_normalization_factor": spec["pb_haircut"],
                    "pb_range": [adopted_pb_low, adopted_pb_high],
                },
                "parameter_basis": (
                    "可持续ROE上限取2028基准ROE并封顶22%，下限按公司项目成熟度、"
                    f"资源覆盖和业务质量将上限乘{spec['pb_haircut']:.0%}。低值使用"
                    "12.5%股权成本/2.0%长期增长，高值使用10.5%/3.0%；区间由"
                    "ROE、股权成本和增长敏感性直接计算，不再设置最低PB托底。"
                ),
                "limitation": (
                    "PB只适合净资产能代表已投产经营资本的主体；权益法利润、未投产"
                    "项目、加工库存或前期减值会削弱解释力。"
                ),
            }
        )
    if spec["fcfe_conversion"] > 0 and all(row["fcfe_rmb_bn"] > 0 for row in base):
        dcf = _fcfe_value(
            [row["fcfe_rmb_bn"] for row in base],
            cost_of_equity=0.115,
            terminal_growth=0.025,
        )
        dcf_low = _fcfe_value(
            [row["fcfe_rmb_bn"] for row in base],
            cost_of_equity=0.125,
            terminal_growth=0.02,
        )
        dcf_high = _fcfe_value(
            [row["fcfe_rmb_bn"] for row in base],
            cost_of_equity=0.105,
            terminal_growth=0.03,
        )
        valuations.append(
            {
                "method": "股权自由现金流",
                "role": "诊断",
                "forecast_year": 2026,
                "low_rmb_bn": round(dcf_low["equity_value_rmb_bn"], 2),
                "high_rmb_bn": round(dcf_high["equity_value_rmb_bn"], 2),
                "inputs": {
                    "fcfe_rmb_bn": [row["fcfe_rmb_bn"] for row in base],
                    "base_cost_of_equity_pct": 11.5,
                    "base_terminal_growth_pct": 2.5,
                    "low_value_cost_of_equity_pct": 12.5,
                    "low_value_terminal_growth_pct": 2.0,
                    "high_value_cost_of_equity_pct": 10.5,
                    "high_value_terminal_growth_pct": 3.0,
                    "base_terminal_value_share_pct": round(
                        dcf["terminal_value_share_pct"], 2
                    ),
                    "low_terminal_value_share_pct": round(
                        dcf_low["terminal_value_share_pct"], 2
                    ),
                    "high_terminal_value_share_pct": round(
                        dcf_high["terminal_value_share_pct"], 2
                    ),
                },
                "parameter_basis": (
                    "下限使用12.5%股权成本/2.0%永续增长，上限使用10.5%/3.0%；"
                    "这是同一组三年FCFE的参数敏感性，不是对单点估值任意加减15%。"
                ),
                "limitation": "资本开支和营运资金预测有限，若终值占比过高则只作参考。",
            }
        )
    if name == "西藏城投":
        valuations.append(
            {
                "method": "条件化项目NAV",
                "role": "未采用",
                "forecast_year": 2028,
                "low_rmb_bn": None,
                "high_rmb_bn": None,
                "inputs": {
                    "company_project_ownership_pct": 41.0,
                    "planned_capacity_10kt": 10.0,
                },
                "parameter_basis": (
                    "已确认国能矿业41%权益和两盐湖规划产能，但尚缺同口径储量采序、"
                    "回收率、建设与维持资本、单位成本、税制和投产时间，无法把规划"
                    "产能可靠地转成项目净现值。"
                ),
                "limitation": (
                    "关闭原4—8亿元条件值：该数值没有完整现金流底稿，也遗漏地产等"
                    "非锂资产，既不能当项目NAV，也不能与公司总市值直接比较。"
                ),
            }
        )

    core_values = [
        value
        for valuation in valuations
        if valuation["role"] == "核心"
        for value in (valuation["low_rmb_bn"], valuation["high_rmb_bn"])
        if value is not None and math.isfinite(value)
    ]
    core_range = (
        {
            "low_rmb_bn": round(min(core_values), 2),
            "high_rmb_bn": round(max(core_values), 2),
            "method": "只使用标记为核心的方法；参考与诊断结果不并入区间",
        }
        if core_values
        else {
            "low_rmb_bn": None,
            "high_rmb_bn": None,
            "method": "没有满足数据门槛的核心估值方法，暂不形成总股权价值区间",
        }
    )
    return {
        "company": name,
        "research_company_id": spec["company_id"],
        "ticker": spec["ticker"],
        "model_type": spec["model_type"],
        "actual_2025": {key: round(value, 4) for key, value in actual.items()},
        "project_and_operating_evidence": spec["source_facts"],
        "assumptions": {
            key: value
            for key, value in spec.items()
            if key
            in {
                "product_volume",
                "resource_share",
                "resource_cost",
                "processing_margin",
                "after_tax_factor",
                "other_profit",
                "corporate_cost",
                "revenue",
                "equity",
                "fcfe_conversion",
                "net_income_override",
                "net_income_override_rationale",
            }
        },
        "formula": {
            "resource_profit": (
                "资源税后利润＝产品销量×资源自给比例×（含税锂价－含税资源成本）"
                "÷1.13×税后归母转换系数"
            ),
            "processing_profit": (
                "加工税后利润＝产品销量×（1－资源自给比例）×单位加工利润"
                "÷1.13×税后归母转换系数"
            ),
            "net_income": (
                "归母净利润＝未商业化项目投产前的公司独立损失路径；"
                "规划产能和试验性产量不计入基准利润"
                if spec.get("net_income_override")
                else "归母净利润＝资源利润＋加工利润＋其他业务利润－公司层成本"
            ),
        },
        "scenarios": scenarios,
        "valuations": valuations,
        "independent_equity_value_range": core_range,
        "valuation_note": spec["valuation_note"],
        "limitations": [
            "权益产量、资源自给率和成本包含研究估算，已在输入中逐年冻结，不是公司指引。",
            "价格情景为含税人民币现货等价，不代表不同品级和长协的实际成交价。",
            "公司层模型只给与输入精度匹配的区间，不能把项目产能等同于当年销售量。",
            *(
                [str(spec["net_income_override_rationale"])]
                if spec.get("net_income_override_rationale")
                else []
            ),
        ],
    }


def build() -> dict[str, Any]:
    workbook_contract = _workbook_contract()
    actuals = _snapshot_actuals()
    lithium_supply = _build_lithium_supply_demand()
    carbonate_supply = _build_carbonate_supply_demand()
    _write_json(MODEL_DIR / "lithium_supply_demand_model_v1.json", lithium_supply)
    _write_json(MODEL_DIR / "carbonate_supply_demand_model_v1.json", carbonate_supply)

    companies = []
    freeze_records = []
    for name, spec in COMPANIES.items():
        company = _company_model(name, spec, actuals[spec["ticker"]])
        freeze_input = {
            "schema_version": "company_financial_model_freeze_input.v1",
            "as_of_date": MODEL_AS_OF,
            "company": name,
            "ticker": spec["ticker"],
            "workbook_contract_sha256": workbook_contract["sha256"],
            "actual_2025": company["actual_2025"],
            "project_and_operating_evidence": company["project_and_operating_evidence"],
            "assumptions": company["assumptions"],
            "valuation_policy": [
                {
                    "method": valuation["method"],
                    "role": valuation["role"],
                    "forecast_year": valuation["forecast_year"],
                    "inputs": valuation.get("inputs") or {},
                    "parameter_basis": valuation.get("parameter_basis") or "",
                    "limitation": valuation.get("limitation") or "",
                }
                for valuation in company["valuations"]
            ],
            "price_scenarios": PRICE_SCENARIOS,
            "formula": company["formula"],
            "external_consensus_read": False,
            "current_market_snapshot_read": False,
        }
        freeze_output = {
            "schema_version": "company_financial_model_freeze_output.v1",
            "as_of_date": MODEL_AS_OF,
            "company": name,
            "ticker": spec["ticker"],
            "scenarios": company["scenarios"],
            "valuations": company["valuations"],
            "independent_equity_value_range": company["independent_equity_value_range"],
        }
        safe_ticker = spec["ticker"].replace(".", "_")
        input_path = FREEZE_DIR / f"{safe_ticker}_input.json"
        output_path = FREEZE_DIR / f"{safe_ticker}_output.json"
        _write_json(input_path, freeze_input)
        _write_json(output_path, freeze_output)
        input_sha = _sha256_file(input_path)
        output_sha = _sha256_file(output_path)
        company["freeze"] = {
            "input_path": input_path.relative_to(ROOT).as_posix(),
            "input_sha256": input_sha,
            "output_path": output_path.relative_to(ROOT).as_posix(),
            "output_sha256": output_sha,
            "frozen_before_external_reconciliation": True,
        }
        freeze_records.append({"company": name, **company["freeze"]})
        companies.append(company)

    payload = {
        "schema_version": "lithium_company_independent_models.v1",
        "as_of_date": MODEL_AS_OF,
        "research_run_ref": "btrack_lithium_and_carbonate_20260727",
        "workbook_contract": workbook_contract,
        "supply_demand_artifacts": {
            "lithium": {
                "path": "cache/lithium_research/models/lithium_supply_demand_model_v1.json",
                "sha256": _sha256_file(
                    MODEL_DIR / "lithium_supply_demand_model_v1.json"
                ),
            },
            "carbonate": {
                "path": "cache/lithium_research/models/carbonate_supply_demand_model_v1.json",
                "sha256": _sha256_file(
                    MODEL_DIR / "carbonate_supply_demand_model_v1.json"
                ),
            },
        },
        "independent_freeze": {
            "external_consensus_read": False,
            "current_market_snapshot_read": False,
            "company_count": len(companies),
            "records": freeze_records,
        },
        "companies": companies,
    }
    payload["content_sha256"] = _sha256(payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    payload = build()
    _write_json(args.output, payload)
    print(
        json.dumps(
            {
                "output": str(args.output),
                "companies": len(payload["companies"]),
                "content_sha256": payload["content_sha256"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
